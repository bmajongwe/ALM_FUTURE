from collections import defaultdict
from datetime import datetime
import logging
from pyexpat.errors import messages
from tkinter.font import Font
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F
import openpyxl
from ALM_APP.Functions.liquidity_gap_utils import calculate_totals, filter_queryset_by_form, get_date_buckets, get_latest_fic_mis_date, prepare_inflow_outflow_data
from ALM_APP.forms import LiquidityGapReportFilterForm, LiquidityGapReportFilterForm_cons
from ALM_APP.models import LiquidityGapResultsBase, LiquidityGapResultsCons, Log
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import Workbook
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden











logger = logging.getLogger(__name__)

@login_required


def liquidity_gap_report_base(request):
    """
    View to display a base Liquidity Gap Report. Reads from session and GET parameters to filter data.
    Produces a multi-currency view of inflows, outflows, net liquidity gaps, and optional drill-downs.
    Logs actions to the Log table but does not integrate audit trails.
    """
    logger.info("Accessed liquidity_gap_report_base view.")
    Log.objects.create(
        function_name='liquidity_gap_report_base',
        log_level='INFO',
        message="Accessed liquidity_gap_report_base view.",
        status='SUCCESS'
    )

    # Step 1: Manage GET parameters and session-based filters
    if request.GET:
        logger.debug(f"Received GET parameters for liquidity_gap_report_base: {request.GET}")
        Log.objects.create(
            function_name='liquidity_gap_report_base',
            log_level='DEBUG',
            message=f"Received GET parameters: {request.GET}",
            status='SUCCESS'
        )

        session_filters = request.session.get('filters_base', {})
        new_params = request.GET.dict()
        merged_filters = {**session_filters, **new_params}
        request.session['filters_base'] = merged_filters
        filters = merged_filters
    else:
        filters = request.session.get('filters_base', {})

    # Step 2: Initialize the form and base queryset
    form = LiquidityGapReportFilterForm(filters or None)
    base_queryset = LiquidityGapResultsBase.objects.all()

    # Step 3: Determine fic_mis_date from form or fallback to the latest
    if form.is_valid():
        fic_mis_date = form.cleaned_data.get('fic_mis_date')
    else:
        fic_mis_date = None

    if not fic_mis_date:
        fic_mis_date = get_latest_fic_mis_date()
        if not fic_mis_date:
            error_message = "No data available for the selected filters."
            logger.warning(error_message)
            Log.objects.create(
                function_name='liquidity_gap_report_base',
                log_level='WARNING',
                message=error_message,
                status='SUCCESS'
            )
            messages.error(request, error_message)
            return render(
                request,
                'ALM_APP/reports/liquidity_gap_report_base.html',
                {'form': form}
            )

    logger.debug(f"Using fic_mis_date='{fic_mis_date}' for liquidity gap report.")
    Log.objects.create(
        function_name='liquidity_gap_report_base',
        log_level='DEBUG',
        message=f"Using fic_mis_date='{fic_mis_date}' for report.",
        status='SUCCESS'
    )

    # Step 4: Retrieve date buckets for the fic_mis_date
    date_buckets = get_date_buckets(fic_mis_date)
    if not date_buckets.exists():
        error_message = "No date buckets available for the selected filters."
        logger.warning(error_message)
        Log.objects.create(
            function_name='liquidity_gap_report_base',
            log_level='WARNING',
            message=error_message,
            status='SUCCESS'
        )
        messages.error(request, error_message)
        return render(
            request,
            'ALM_APP/reports/liquidity_gap_report_base.html',
            {'form': form}
        )

    # Step 5: Read additional drill-down and currency params
    drill_down_product = request.GET.get('drill_down_product')
    drill_down_splits = request.GET.get('drill_down_splits')
    selected_currency = request.GET.get('selected_currency', None)

    logger.debug(
        f"Drill-down params: product='{drill_down_product}', splits='{drill_down_splits}', currency='{selected_currency}'."
    )
    Log.objects.create(
        function_name='liquidity_gap_report_base',
        log_level='DEBUG',
        message=(
            f"Drill-down params: product='{drill_down_product}', "
            f"splits='{drill_down_splits}', currency='{selected_currency}'."
        ),
        status='SUCCESS'
    )

    # Step 6: Filter the base queryset by form + fic_mis_date
    base_queryset = filter_queryset_by_form(form, base_queryset).filter(fic_mis_date=fic_mis_date)
    logger.debug(f"Base queryset filtered with {base_queryset.count()} records remaining.")
    Log.objects.create(
        function_name='liquidity_gap_report_base',
        log_level='DEBUG',
        message=f"Base queryset filtered to {base_queryset.count()} records.",
        status='SUCCESS'
    )

    # Step 7: Prepare currency-specific data structure
    from collections import defaultdict
    currency_data = defaultdict(lambda: {
        'inflow_data': {},
        'outflow_data': {},
        'net_liquidity_gap': {},
        'net_gap_percentage': {},
        'cumulative_gap': {},
        'first_inflow_product': None,
        'remaining_inflow_data': {},
        'first_outflow_product': None,
        'remaining_outflow_data': {},
        'aggregated_product_details': None,
        'aggregated_split_details': None,
    })

    # Determine which currencies to iterate over
    if selected_currency:
        currencies = [selected_currency]
    else:
        currencies = base_queryset.values_list('v_ccy_code', flat=True).distinct()
    logger.info(f"Processing currencies: {list(currencies)}")
    Log.objects.create(
        function_name='liquidity_gap_report_base',
        log_level='INFO',
        message=f"Processing currencies: {list(currencies)}",
        status='SUCCESS'
    )

    # Step 8: Build currency-specific data
    for currency in currencies:
        currency_queryset = base_queryset.filter(v_ccy_code=currency)
        local_drill_down_qs = currency_queryset

        if drill_down_product:
            local_drill_down_qs = local_drill_down_qs.filter(v_prod_type=drill_down_product)
        if drill_down_splits:
            local_drill_down_qs = local_drill_down_qs.filter(v_product_name=drill_down_splits)

        # Aggregated drill-down logic
        local_aggregated_product_details = None
        local_aggregated_split_details = None

        # If drilling down by splits
        if drill_down_splits:
            drill_down_splits_details = list(
                local_drill_down_qs
                .values('v_product_splits', 'bucket_number')
                .annotate(
                    inflows_total=Sum('inflows'),
                    outflows_total=Sum('outflows'),
                    total=Sum(F('inflows') - F('outflows'))
                )
            )
            grouped_split_data = defaultdict(lambda: {b['bucket_number']: 0 for b in date_buckets})
            for detail in drill_down_splits_details:
                split_name = detail.get('v_product_splits')
                bucket_number = detail['bucket_number']
                grouped_split_data[split_name][bucket_number] += detail.get('total', 0)

            local_aggregated_split_details = [
                {
                    'v_product_splits': split_name,
                    'buckets': buckets,
                    'total': sum(buckets.values())
                }
                for split_name, buckets in grouped_split_data.items()
            ]

        # If drilling down by product
        elif drill_down_product:
            drill_down_details = list(
                local_drill_down_qs
                .values('v_product_name', 'bucket_number')
                .annotate(
                    inflows_total=Sum('inflows'),
                    outflows_total=Sum('outflows'),
                    total=Sum(F('inflows') - F('outflows'))
                )
            )
            grouped_data = defaultdict(lambda: {b['bucket_number']: 0 for b in date_buckets})
            for detail in drill_down_details:
                product_name = detail.get('v_product_name')
                bucket_number = detail['bucket_number']
                grouped_data[product_name][bucket_number] += detail.get('total', 0)

            local_aggregated_product_details = [
                {
                    'v_product_name': product_name,
                    'buckets': buckets,
                    'total': sum(buckets.values())
                }
                for product_name, buckets in grouped_data.items()
            ]

        # Prepare inflow/outflow data
        inflow_data, outflow_data = prepare_inflow_outflow_data(local_drill_down_qs)
        net_liquidity_gap, net_gap_percentage, cumulative_gap = calculate_totals(
            date_buckets, inflow_data, outflow_data
        )

        # Compute total for each product in inflow_data
        for product, buckets in inflow_data.items():
            inflow_data[product]['total'] = sum(
                buckets.get(b['bucket_number'], 0) for b in date_buckets
            )
        # Compute total for each product in outflow_data
        for product, buckets in outflow_data.items():
            outflow_data[product]['total'] = sum(
                buckets.get(b['bucket_number'], 0) for b in date_buckets
            )

        # Compute total for net_liquidity_gap
        net_liquidity_gap['total'] = sum(
            net_liquidity_gap.get(b['bucket_number'], 0) for b in date_buckets
        )
        # Compute total for net_gap_percentage
        net_gap_percentage['total'] = (
            sum(net_gap_percentage.get(b['bucket_number'], 0) for b in date_buckets) / len(date_buckets)
        ) if len(date_buckets) > 0 else 0

        # Compute total for cumulative_gap
        last_bucket = date_buckets.last()
        cumulative_gap['total'] = cumulative_gap.get(last_bucket['bucket_number'], 0) if last_bucket else 0

        # Identify first/remaining inflow product
        if inflow_data:
            first_inflow_product_key = list(inflow_data.keys())[0]
            first_inflow_product = (first_inflow_product_key, inflow_data[first_inflow_product_key])
            remaining_inflow_data = inflow_data.copy()
            remaining_inflow_data.pop(first_inflow_product_key, None)
        else:
            first_inflow_product = None
            remaining_inflow_data = {}

        # Identify first/remaining outflow product
        if outflow_data:
            first_outflow_product_key = list(outflow_data.keys())[0]
            first_outflow_product = (first_outflow_product_key, outflow_data[first_outflow_product_key])
            remaining_outflow_data = outflow_data.copy()
            remaining_outflow_data.pop(first_outflow_product_key, None)
        else:
            first_outflow_product = None
            remaining_outflow_data = {}

        # Update currency_data structure
        currency_data[currency].update({
            'inflow_data': inflow_data,
            'outflow_data': outflow_data,
            'first_inflow_product': first_inflow_product,
            'remaining_inflow_data': remaining_inflow_data,
            'first_outflow_product': first_outflow_product,
            'remaining_outflow_data': remaining_outflow_data,
            'net_liquidity_gap': net_liquidity_gap,
            'net_gap_percentage': net_gap_percentage,
            'cumulative_gap': cumulative_gap,
            'aggregated_product_details': local_aggregated_product_details,
            'aggregated_split_details': local_aggregated_split_details,
        })

        logger.debug(f"Processed currency='{currency}' in liquidity gap report with {len(inflow_data)} inflow items and {len(outflow_data)} outflow items.")
        Log.objects.create(
            function_name='liquidity_gap_report_base',
            log_level='DEBUG',
            message=(
                f"Processed currency='{currency}' with {len(inflow_data)} inflow items "
                f"and {len(outflow_data)} outflow items."
            ),
            status='SUCCESS'
        )

    # Prepare context for rendering
    context = {
        'form': form,
        'fic_mis_date': fic_mis_date,
        'date_buckets': date_buckets,
        'currency_data': dict(currency_data),
        'total_columns': len(date_buckets) + 3,
        'drill_down_product': drill_down_product,
        'drill_down_splits': drill_down_splits,
        'selected_currency': selected_currency,
    }

    logger.info("Finished building data for liquidity_gap_report_base. Rendering template.")
    Log.objects.create(
        function_name='liquidity_gap_report_base',
        log_level='INFO',
        message="Finished building data for liquidity_gap_report_base. Rendering template.",
        status='SUCCESS'
    )

    return render(request, 'ALM_APP/reports/liquidity_gap_report_base.html', context)


#######################



########################################################################################################################


logger = logging.getLogger(__name__)
@login_required
def liquidity_gap_report_cons(request):
    """
    View to display a consolidated Liquidity Gap Report. Reads from session and GET parameters
    to filter data. Produces a multi-currency view of inflows, outflows, net liquidity gaps,
    and optional drill-downs. Logs actions to the Log table but does not integrate audit trails.
    """
    logger.info("Accessed liquidity_gap_report_cons view.")
    Log.objects.create(
        function_name='liquidity_gap_report_cons',
        log_level='INFO',
        message="Accessed liquidity_gap_report_cons view.",
        status='SUCCESS'
    )

    # Step 1: Manage GET parameters and session-based filters
    if request.GET:
        logger.debug(f"Received GET parameters for liquidity_gap_report_cons: {request.GET}")
        Log.objects.create(
            function_name='liquidity_gap_report_cons',
            log_level='DEBUG',
            message=f"Received GET parameters: {request.GET}",
            status='SUCCESS'
        )

        session_filters = request.session.get('filters_cons', {})
        new_params = request.GET.dict()
        merged_filters = {**session_filters, **new_params}
        request.session['filters_cons'] = merged_filters
        filters = merged_filters
    else:
        filters = request.session.get('filters_cons', {})

    # Step 2: Initialize the form and cons_queryset
    form = LiquidityGapReportFilterForm_cons(filters or None)
    cons_queryset = LiquidityGapResultsCons.objects.all()
    logger.debug("Initialized LiquidityGapReportFilterForm_cons and base cons_queryset.")
    Log.objects.create(
        function_name='liquidity_gap_report_cons',
        log_level='DEBUG',
        message="Initialized LiquidityGapReportFilterForm_cons and base cons_queryset.",
        status='SUCCESS'
    )

    # Step 3: Determine fic_mis_date from the form or fallback to the latest
    if form.is_valid():
        fic_mis_date = form.cleaned_data.get('fic_mis_date')
    else:
        fic_mis_date = None

    if not fic_mis_date:
        fic_mis_date = get_latest_fic_mis_date()
        if not fic_mis_date:
            error_message = "No data available for the selected filters."
            logger.warning(error_message)
            Log.objects.create(
                function_name='liquidity_gap_report_cons',
                log_level='WARNING',
                message=error_message,
                status='SUCCESS'
            )
            messages.error(request, error_message)
            return render(request, 'ALM_APP/reports/liquidity_gap_report_cons.html', {'form': form})

    logger.debug(f"Using fic_mis_date='{fic_mis_date}' for liquidity gap report (consolidated).")
    Log.objects.create(
        function_name='liquidity_gap_report_cons',
        log_level='DEBUG',
        message=f"Using fic_mis_date='{fic_mis_date}' for report (consolidated).",
        status='SUCCESS'
    )

    # Step 4: Retrieve date buckets for fic_mis_date
    date_buckets = get_date_buckets(fic_mis_date)
    if not date_buckets.exists():
        error_message = "No date buckets available for the selected filters."
        logger.warning(error_message)
        Log.objects.create(
            function_name='liquidity_gap_report_cons',
            log_level='WARNING',
            message=error_message,
            status='SUCCESS'
        )
        messages.error(request, error_message)
        return render(request, 'ALM_APP/reports/liquidity_gap_report_cons.html', {'form': form})

    # Step 5: Drill-down parameters
    drill_down_product_cons = request.GET.get('drill_down_product_cons')
    drill_down_splits_cons = request.GET.get('drill_down_splits_cons')
    logger.debug(
        f"Drill-down params: product='{drill_down_product_cons}', splits='{drill_down_splits_cons}'."
    )
    Log.objects.create(
        function_name='liquidity_gap_report_cons',
        log_level='DEBUG',
        message=(
            f"Drill-down params: product='{drill_down_product_cons}', "
            f"splits='{drill_down_splits_cons}'."
        ),
        status='SUCCESS'
    )

    # Step 6: Filter the queryset
    cons_queryset = filter_queryset_by_form(form, cons_queryset).filter(fic_mis_date=fic_mis_date)
    logger.debug(f"Filtered cons_queryset to {cons_queryset.count()} records.")
    Log.objects.create(
        function_name='liquidity_gap_report_cons',
        log_level='DEBUG',
        message=f"Filtered cons_queryset to {cons_queryset.count()} records.",
        status='SUCCESS'
    )

    drill_down_details_cons = None
    drill_down_splits_details_cons = None
    aggregated_product_details_cons = None
    aggregated_split_details_cons = None

    # Step 7: Handle splits drill-down if specified
    if drill_down_splits_cons:
        drill_down_splits_details_cons = list(
            cons_queryset.filter(v_product_name=drill_down_splits_cons)
            .values('v_product_splits', 'bucket_number')
            .annotate(
                inflows_total=Sum('inflows'),
                outflows_total=Sum('outflows'),
                total=Sum(F('inflows') - F('outflows'))
            )
        )
        grouped_split_data_cons = defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})
        for detail in drill_down_splits_details_cons:
            split_name = detail.get('v_product_splits')
            bucket_number = detail['bucket_number']
            grouped_split_data_cons[split_name][bucket_number] += detail.get('total', 0)

        aggregated_split_details_cons = [
            {
                'v_product_splits': split_name,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for split_name, buckets in grouped_split_data_cons.items()
        ]

    # Step 8: Handle product drill-down if specified
    elif drill_down_product_cons:
        drill_down_details_cons = list(
            cons_queryset.filter(v_prod_type=drill_down_product_cons)
            .values('v_product_name', 'bucket_number')
            .annotate(
                inflows_total=Sum('inflows'),
                outflows_total=Sum('outflows'),
                total=Sum(F('inflows') - F('outflows'))
            )
        )
        grouped_data_cons = defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})
        for detail in drill_down_details_cons:
            product_name = detail.get('v_product_name')
            bucket_number = detail['bucket_number']
            grouped_data_cons[product_name][bucket_number] += detail.get('total', 0)

        aggregated_product_details_cons = [
            {
                'v_product_name': product_name,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for product_name, buckets in grouped_data_cons.items()
        ]

    # Step 9: Prepare inflow and outflow data
    cons_inflow_data, cons_outflow_data = prepare_inflow_outflow_data(cons_queryset)
    cons_net_liquidity_gap, cons_net_gap_percentage, cons_cumulative_gap = calculate_totals(
        date_buckets, cons_inflow_data, cons_outflow_data
    )

    # Extract first and remaining inflow data
    if cons_inflow_data:
        cons_first_inflow_product = list(cons_inflow_data.items())[0]
        cons_remaining_inflow_data = cons_inflow_data.copy()
        cons_remaining_inflow_data.pop(cons_first_inflow_product[0], None)
    else:
        cons_first_inflow_product = None
        cons_remaining_inflow_data = {}

    # Extract first and remaining outflow data
    if cons_outflow_data:
        cons_first_outflow_product = list(cons_outflow_data.items())[0]
        cons_remaining_outflow_data = cons_outflow_data.copy()
        cons_remaining_outflow_data.pop(cons_first_outflow_product[0], None)
    else:
        cons_first_outflow_product = None
        cons_remaining_outflow_data = {}

    cons_data = {
        'inflow_data': cons_inflow_data,
        'outflow_data': cons_outflow_data,
        'first_inflow_product': cons_first_inflow_product,
        'remaining_inflow_data': cons_remaining_inflow_data,
        'first_outflow_product': cons_first_outflow_product,
        'remaining_outflow_data': cons_remaining_outflow_data,
        'net_liquidity_gap': cons_net_liquidity_gap,
        'net_gap_percentage': cons_net_gap_percentage,
        'cumulative_gap': cons_cumulative_gap,
    }

    logger.debug(f"Prepared consolidated data for liquidity gap: {cons_data}")
    # Do not log the entire structure if it's too large, for brevity we do minimal logs
    Log.objects.create(
        function_name='liquidity_gap_report_cons',
        log_level='DEBUG',
        message=f"Prepared consolidated data for liquidity gap (keys: {list(cons_data.keys())}).",
        status='SUCCESS'
    )

    # Optional sub-drill data
    drill_cons_data = None
    if drill_down_product_cons or drill_down_splits_cons:
        if drill_down_product_cons:
            drill_queryset = cons_queryset.filter(v_prod_type=drill_down_product_cons)
        elif drill_down_splits_cons:
            drill_queryset = cons_queryset.filter(v_product_name=drill_down_splits_cons)

        drill_inflow_data, drill_outflow_data = prepare_inflow_outflow_data(drill_queryset)
        drill_net_liquidity_gap, drill_net_gap_percentage, drill_cumulative_gap = calculate_totals(
            date_buckets, drill_inflow_data, drill_outflow_data
        )

        drill_cons_data = {
            'inflow_data': drill_inflow_data,
            'outflow_data': drill_outflow_data,
            'net_liquidity_gap': drill_net_liquidity_gap,
            'net_gap_percentage': drill_net_gap_percentage,
            'cumulative_gap': drill_cumulative_gap,
        }
        logger.debug("Prepared drill-down data for consolidated liquidity gap report.")
        Log.objects.create(
            function_name='liquidity_gap_report_cons',
            log_level='DEBUG',
            message="Prepared drill-down data for consolidated liquidity gap report.",
            status='SUCCESS'
        )

    # Extract distinct currencies
    currency_data = cons_queryset.values_list('v_ccy_code', flat=True).distinct()
    logger.debug(f"Extracted currency data for consolidated report: {list(currency_data)}")
    Log.objects.create(
        function_name='liquidity_gap_report_cons',
        log_level='DEBUG',
        message=f"Extracted currency data: {list(currency_data)}",
        status='SUCCESS'
    )

    context = {
        'form': form,
        'fic_mis_date': fic_mis_date,
        'date_buckets': date_buckets,
        'currency_data': currency_data,
        'cons_data': cons_data,
        'drill_cons_data': drill_cons_data,
        'total_columns': len(date_buckets) + 3,
        'drill_down_product_cons': drill_down_product_cons,
        'drill_down_splits_cons': drill_down_splits_cons,
        'aggregated_product_details_cons': aggregated_product_details_cons,
        'aggregated_split_details_cons': aggregated_split_details_cons,
    }

    logger.info("Rendering liquidity_gap_report_cons template with consolidated data.")
    Log.objects.create(
        function_name='liquidity_gap_report_cons',
        log_level='INFO',
        message="Rendering liquidity_gap_report_cons template with consolidated data.",
        status='SUCCESS'
    )

    return render(request, 'ALM_APP/reports/liquidity_gap_report_cons.html', context)




@login_required

def export_liquidity_gap_to_excel(request):
    form = LiquidityGapReportFilterForm(request.GET or None)

    # Parse fic_mis_date from request
    raw_fic_mis_date = request.GET.get('fic_mis_date')
    fic_mis_date = parse_date(raw_fic_mis_date)

    if not fic_mis_date:
        return HttpResponseBadRequest(
            "Invalid date format. Supported formats: 'Aug. 31, 2024', 'August 31, 2024', "
            "'2024-08-31', '31-08-2024', '31/08/2024', '08/31/2024'."
        )

    # Base queryset filtered by fic_mis_date
    queryset = LiquidityGapResultsBase.objects.filter(fic_mis_date=fic_mis_date)
    queryset = filter_queryset_by_form(form, queryset)

    # Handle drill-down parameters
    drill_down_product = request.GET.get('drill_down_product')
    drill_down_splits = request.GET.get('drill_down_splits')

    if drill_down_product:
        queryset = queryset.filter(v_prod_type=drill_down_product)
    if drill_down_splits:
        queryset = queryset.filter(v_product_name=drill_down_splits)

    # Get date buckets
    date_buckets = get_date_buckets(fic_mis_date)
    if not date_buckets.exists():
        return HttpResponseBadRequest("No date buckets available for the selected date.")

    # Prepare headers based on drill-down level
    headers = ["Account Type", "Product"]
    if drill_down_splits:
        headers += ["v_product_name", "v_product_splits"]
    elif drill_down_product:
        headers += ["v_product_name"]

    headers += [
        f"{bucket['bucket_start_date'].strftime('%d-%b-%Y')} to {bucket['bucket_end_date'].strftime('%d-%b-%Y')}"
        for bucket in date_buckets
    ] + ["Total"]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"), right=Side(border_style="thin"),
        top=Side(border_style="thin"), bottom=Side(border_style="thin")
    )
    alignment_center = Alignment(horizontal="center", vertical="center")
    heading_font = Font(bold=True, size=14)

    # Overview scenario: separate sheets per currency
    if not drill_down_product and not drill_down_splits:
        # Group data by currency (v_ccy_code) and product type
        details = queryset.values('v_ccy_code', 'v_prod_type', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        currency_grouped = defaultdict(lambda: defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets}))
        for detail in details:
            currency = detail['v_ccy_code']  # Use v_ccy_code instead of currency
            prod_type = detail['v_prod_type']
            bucket_number = detail['bucket_number']
            currency_grouped[currency][prod_type][bucket_number] += detail.get('total', 0)

        # Initialize workbook and remove the default sheet
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        # Create a sheet for each currency
        for currency, products in currency_grouped.items():
            sheet = workbook.create_sheet(title=str(currency))
            total_columns = len(headers)

            # Dynamic heading for this currency sheet
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
            heading_cell = sheet.cell(row=1, column=1)
            heading_cell.value = f"Overview Report for {currency} on {fic_mis_date}"
            heading_cell.font = heading_font
            heading_cell.alignment = alignment_center

            # Write headers in row 2
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=2, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = border

            # Write aggregated data rows starting from row 3
            row_num = 3
            for prod_type, buckets in products.items():
                account_type = "Total Inflow" if sum(buckets.values()) >= 0 else "Total Outflow"
                row_prefix = [account_type, prod_type]
                row_data = row_prefix + [
                    buckets.get(bucket['bucket_number'], 0) for bucket in date_buckets
                ] + [sum(buckets.values())]

                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = alignment_center
                    cell.border = border
                row_num += 1

            # Compute summary rows for the current currency
            currency_qs = queryset.filter(v_ccy_code=currency)  # Filter by current currency
            inflow_data_cur, outflow_data_cur = prepare_inflow_outflow_data(currency_qs)
            net_liquidity_gap_cur, net_gap_percentage_cur, cumulative_gap_cur = calculate_totals(
                date_buckets, inflow_data_cur, outflow_data_cur
            )

            total_outflows_cur = sum(outflow_data_cur.get(bucket['bucket_number'], 0) for bucket in date_buckets)
            net_liquidity_gap_total_cur = sum(net_liquidity_gap_cur.get(bucket['bucket_number'], 0) for bucket in date_buckets)
            cumulative_gap_total_cur = sum(net_liquidity_gap_cur.get(bucket['bucket_number'], 0) for bucket in date_buckets)
            net_gap_percentage_total_cur = (net_liquidity_gap_total_cur / total_outflows_cur * 100) if total_outflows_cur else 0

            prefix_count = 2
            summary_rows = [
                ("Net Liquidity Gap", net_liquidity_gap_cur, net_liquidity_gap_total_cur),
                ("Net Gap as % of Total Outflows", net_gap_percentage_cur, net_gap_percentage_total_cur),
                ("Cumulative Gap", cumulative_gap_cur, cumulative_gap_total_cur),
            ]
            for label, summary_data, summary_total in summary_rows:
                row = [label] + [""] * (prefix_count - 1)
                row += [summary_data.get(bucket['bucket_number'], 0) for bucket in date_buckets]
                row.append(summary_total)

                for col_num, value in enumerate(row, 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.font = Font(bold=True)
                    cell.alignment = alignment_center
                    cell.border = border
                row_num += 1

            # Auto-adjust column widths for this sheet
            for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                max_length = 0
                for cell in col:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max_length + 2

        # Save and return the workbook for overview scenario
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="LiquidityGapReport_Overview_{fic_mis_date}.xlsx"'
        workbook.save(response)
        return response

    # Non-overview (drill-down) scenarios
    grouped_data = defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})

    if drill_down_splits:
        details = queryset.values('v_product_splits', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            split_name = detail['v_product_splits']
            bucket_number = detail['bucket_number']
            grouped_data[split_name][bucket_number] += detail.get('total', 0)
        aggregated_data = [
            {'identifier': split_name, 'buckets': buckets, 'total': sum(buckets.values())}
            for split_name, buckets in grouped_data.items()
        ]
    elif drill_down_product:
        details = queryset.values('v_product_name', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            product_name = detail['v_product_name']
            bucket_number = detail['bucket_number']
            grouped_data[product_name][bucket_number] += detail.get('total', 0)
        aggregated_data = [
            {'identifier': product_name, 'buckets': buckets, 'total': sum(buckets.values())}
            for product_name, buckets in grouped_data.items()
        ]
    else:
        details = queryset.values('v_prod_type', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            prod_type = detail['v_prod_type']
            bucket_number = detail['bucket_number']
            grouped_data[prod_type][bucket_number] += detail.get('total', 0)
        aggregated_data = [
            {'identifier': prod_type, 'buckets': buckets, 'total': sum(buckets.values())}
            for prod_type, buckets in grouped_data.items()
        ]

    # Calculate totals for summary rows for non-overview
    inflow_data, outflow_data = prepare_inflow_outflow_data(queryset)
    net_liquidity_gap, net_gap_percentage, cumulative_gap = calculate_totals(
        date_buckets, inflow_data, outflow_data
    )

    total_outflows = sum(outflow_data.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    net_liquidity_gap_total = sum(net_liquidity_gap.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    cumulative_gap_total = sum(cumulative_gap.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    net_gap_percentage_total = (net_liquidity_gap_total / total_outflows * 100) if total_outflows else 0

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Liquidity Gap Report"

    # Determine dynamic heading based on drill-down level
    if drill_down_splits:
        heading_text = f"Drill-Down Report for Split: {drill_down_splits} on {fic_mis_date}"
    elif drill_down_product:
        heading_text = f"Drill-Down Report for Product: {drill_down_product} on {fic_mis_date}"
    else:
        heading_text = f"Overview Report for {fic_mis_date}"

    total_columns = len(headers)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    heading_cell = sheet.cell(row=1, column=1)
    heading_cell.value = heading_text
    heading_cell.font = heading_font
    heading_cell.alignment = alignment_center

    # Write headers for non-overview
    sheet.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = border

    # Write aggregated data rows for non-overview
    row_num = 3
    for item in aggregated_data:
        if drill_down_splits:
            row_prefix = ["Total Flows", drill_down_product, drill_down_splits, item['identifier']]
        elif drill_down_product:
            row_prefix = ["Total Flows", drill_down_product, item['identifier']]
        else:
            account_type = "Total Inflow" if item['total'] >= 0 else "Total Outflow"
            row_prefix = [account_type, item['identifier']]

        row = row_prefix + [
            item['buckets'].get(bucket['bucket_number'], 0) for bucket in date_buckets
        ] + [item['total']]

        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = alignment_center
            cell.border = border
        row_num += 1

    if drill_down_splits:
        prefix_count = 4
    elif drill_down_product:
        prefix_count = 3
    else:
        prefix_count = 2

    summary_rows = [
        ("Net Liquidity Gap", net_liquidity_gap, net_liquidity_gap_total),
        ("Net Gap as % of Total Outflows", net_gap_percentage, net_gap_percentage_total),
        ("Cumulative Gap", cumulative_gap, cumulative_gap_total),
    ]

    for label, summary_data, summary_total in summary_rows:
        row = [label] + [""] * (prefix_count - 1)
        row += [summary_data.get(bucket['bucket_number'], 0) for bucket in date_buckets]
        row.append(summary_total)

        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(bold=True)
            cell.alignment = alignment_center
            cell.border = border
        row_num += 1

    for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_length = 0
        for cell in col:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="LiquidityGapReport_{fic_mis_date}.xlsx"'
    workbook.save(response)
    return response







########################################################################################################################################################################
def parse_date(date_str):
    """
    Parse the date string to handle multiple formats.
    """
    date_formats = [
        "%b. %d, %Y",    # 'Aug. 31, 2024'
        "%B %d, %Y",     # 'August 31, 2024'
        "%Y-%m-%d",      # '2024-08-31'
        "%d-%m-%Y",      # '31-08-2024'
        "%d/%m/%Y",      # '31/08/2024'
        "%m/%d/%Y"       # '08/31/2024'
    ]
    for date_format in date_formats:
        try:
            return datetime.strptime(date_str, date_format).date()
        except ValueError:
            continue
    return None




#################################################################################################################################


@login_required
def export_liquidity_gap_cons_to_excel(request):

    form = LiquidityGapReportFilterForm_cons(request.GET or None)

    # Parse fic_mis_date from request
    raw_fic_mis_date = request.GET.get('fic_mis_date')
    fic_mis_date = parse_date(raw_fic_mis_date)

    if not fic_mis_date:
        return HttpResponseBadRequest(
            "Invalid date format. Supported formats: 'Aug. 31, 2024', 'August 31, 2024', "
            "'2024-08-31', '31-08-2024', '31/08/2024', '08/31/2024'."
        )

    # Base queryset filtered by fic_mis_date
    queryset = LiquidityGapResultsCons.objects.filter(fic_mis_date=fic_mis_date)
    queryset = filter_queryset_by_form(form, queryset)

    # Handle drill-down parameters
    drill_down_product_cons = request.GET.get('drill_down_product_cons')
    drill_down_splits_cons = request.GET.get('drill_down_splits_cons')

    if drill_down_product_cons:
        queryset = queryset.filter(v_prod_type=drill_down_product_cons)
    if drill_down_splits_cons:
        queryset = queryset.filter(v_product_name=drill_down_splits_cons)

    # Get date buckets
    date_buckets = get_date_buckets(fic_mis_date)
    if not date_buckets.exists():
        return HttpResponseBadRequest("No date buckets available for the selected date.")

    # Prepare headers based on drill-down level
    headers = ["Account Type", "Product"]
    if drill_down_splits_cons:
        headers += ["v_product_name", "v_product_splits"]
    elif drill_down_product_cons:
        headers += ["v_product_name"]

    # Extend headers with bucket date ranges and Total column
    headers += [
        f"{bucket['bucket_start_date'].strftime('%d-%b-%Y')} to {bucket['bucket_end_date'].strftime('%d-%b-%Y')}"
        for bucket in date_buckets
    ] + ["Total"]

    # Prepare data for the current drill-down level
    grouped_data = defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})

    if drill_down_splits_cons:
        details = queryset.values('v_product_splits', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            split_name = detail['v_product_splits']
            bucket_number = detail['bucket_number']
            grouped_data[split_name][bucket_number] += detail.get('total', 0)

        aggregated_data = [
            {
                'identifier': split_name,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for split_name, buckets in grouped_data.items()
        ]
    elif drill_down_product_cons:
        details = queryset.values('v_product_name', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            product_name = detail['v_product_name']
            bucket_number = detail['bucket_number']
            grouped_data[product_name][bucket_number] += detail.get('total', 0)

        aggregated_data = [
            {
                'identifier': product_name,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for product_name, buckets in grouped_data.items()
        ]
    else:
        # Overview level: Group by v_prod_type when no drill-down filters are applied
        details = queryset.values('v_prod_type', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            prod_type = detail['v_prod_type']
            bucket_number = detail['bucket_number']
            grouped_data[prod_type][bucket_number] += detail.get('total', 0)

        aggregated_data = [
            {
                'identifier': prod_type,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for prod_type, buckets in grouped_data.items()
        ]

    # Calculate totals for summary rows
    inflow_data, outflow_data = prepare_inflow_outflow_data(queryset)
    net_liquidity_gap, net_gap_percentage, cumulative_gap = calculate_totals(
        date_buckets, inflow_data, outflow_data
    )

    total_outflows = sum(outflow_data.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    net_liquidity_gap_total = sum(net_liquidity_gap.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    cumulative_gap_total = sum(net_liquidity_gap.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    if total_outflows:
        net_gap_percentage_total = (net_liquidity_gap_total / total_outflows) * 100
    else:
        net_gap_percentage_total = 0

    # Initialize workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Liquidity Gap Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    alignment_center = Alignment(horizontal="center", vertical="center")
    heading_font = Font(bold=True, size=14)

    # Determine dynamic heading based on drill-down level
    if drill_down_splits_cons:
        heading_text = f"Drill-Down Report for Split: {drill_down_splits_cons} on {fic_mis_date}"
    elif drill_down_product_cons:
        heading_text = f"Drill-Down Report for Product: {drill_down_product_cons} on {fic_mis_date}"
    else:
        heading_text = f"Overview Report for {fic_mis_date}"

    total_columns = len(headers)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    heading_cell = sheet.cell(row=1, column=1)
    heading_cell.value = heading_text
    heading_cell.font = heading_font
    heading_cell.alignment = alignment_center

    # Write headers
    sheet.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = border

    # Write aggregated data rows
    row_num = 3
    for item in aggregated_data:
        if drill_down_splits_cons:
            row_prefix = ["Total Flows", drill_down_product_cons, drill_down_splits_cons, item['identifier']]
        elif drill_down_product_cons:
            row_prefix = ["Total Flows", drill_down_product_cons, item['identifier']]
        else:
            # For overview, determine account type based on total and use v_prod_type as product
            account_type = "Total Inflow" if item['total'] >= 0 else "Total Outflow"
            row_prefix = [account_type, item['identifier']]

        row = row_prefix + [
            item['buckets'].get(bucket['bucket_number'], 0) for bucket in date_buckets
        ] + [item['total']]

        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = alignment_center
            cell.border = border
        row_num += 1

    # Determine fixed prefix column count based on drill-down level
    if drill_down_splits_cons:
        prefix_count = 4
    elif drill_down_product_cons:
        prefix_count = 3
    else:
        prefix_count = 2

    summary_rows = [
        ("Net Liquidity Gap", net_liquidity_gap, net_liquidity_gap_total),
        ("Net Gap as % of Total Outflows", net_gap_percentage, net_gap_percentage_total),
        ("Cumulative Gap", cumulative_gap, cumulative_gap_total),
    ]

    # Write summary rows aligned with the rest of the data
    for label, summary_data, summary_total in summary_rows:
        row = [label] + [""] * (prefix_count - 1)
        row += [summary_data.get(bucket['bucket_number'], 0) for bucket in date_buckets]
        row.append(summary_total)

        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(bold=True)
            cell.alignment = alignment_center
            cell.border = border
        row_num += 1

    # Auto-adjust column widths
    for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_length = 0
        for cell in col:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="LiquidityGapReport_Cons_{fic_mis_date}.xlsx"'
    workbook.save(response)
    return response






