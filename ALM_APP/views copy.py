from .models import LiquidityGapResultsCons
from .Functions.liquidity_gap_utils import filter_queryset_by_form, get_date_buckets, prepare_inflow_outflow_data, calculate_totals
from .models import LiquidityGapResultsBase
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import Workbook
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.db.models import Sum, F
from django.contrib import messages
from .Functions.liquidity_gap_utils import *
from .forms import LiquidityGapReportFilterForm
from .models import LiquidityGapResultsBase, LiquidityGapResultsCons
from django.shortcuts import render
from collections import defaultdict
import os
from django.views.generic.detail import DetailView
from django.urls import reverse_lazy
from django.views import View
import openpyxl
import pandas as pd
from decimal import Decimal
import csv
import traceback
from django.contrib import messages  # Import messages framework
from django.shortcuts import get_object_or_404, render, redirect
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.conf import settings


from .Functions.data import *

from .Functions.Operations import *
from .forms import *
from datetime import datetime
from django.http import HttpResponse
from .Functions.Aggregated_Prod_Cashflow_Base import *
from .Functions.populate_liquidity_gap_results_base import *
from .Functions.ldn_update import *
from .Functions.aggregate_cashflows import *
from .Functions.Aggregated_Acc_level_cashflows import *
from .Functions.behavioral_pattern_utils import define_behavioral_pattern_from_form_data, delete_behavioral_pattern_by_id, update_behavioral_pattern_from_form_data
from .Functions.time_bucket_utils import define_time_bucket_from_form_data, update_time_bucket_from_form_data, delete_time_bucket_by_id
from .Functions.populate_dim import populate_dim_product
from .Functions.Dim_dates import *
from .Functions.product_filter_utils import *
from .Functions.process_utils import *
from .Functions.cashflow import *

from datetime import datetime, timedelta
from django.shortcuts import render, redirect
from .models import TimeBuckets, TimeBucketDefinition, product_level_cashflows


@login_required

def create_behavioral_pattern(request):
    # Fetch the existing Time Bucket Definition (assuming only one is allowed)
    try:
        time_bucket = TimeBucketDefinition.objects.first()
        bucket_entries = TimeBuckets.objects.filter(definition=time_bucket).order_by('serial_number')
    except TimeBucketDefinition.DoesNotExist:
        messages.error(request, "No Time Bucket Definition exists. Please create one first.")
        return redirect('time_bucket_list')

    # Fetch distinct product types from Ldn_Product_Master
    product_types = Ldn_Product_Master.objects.values_list('v_prod_type', flat=True).distinct()

    if request.method == 'POST':
        # Call the function to process the form data
        result = define_behavioral_pattern_from_form_data(request)

        # Check if there is an error
        if 'error' in result:
            messages.error(request, result['error'])

            # Return the form with existing data to repopulate the form fields
            return render(request, 'ALM_APP/behavioral/create_behavioral_pattern.html', {
                'v_prod_type': request.POST.get('v_prod_type'),
                'description': request.POST.get('description'),
                'percentages': request.POST.getlist('percentage[]'),
                'bucket_entries': bucket_entries,  # Include time bucket data for prepopulating
                'product_types': product_types,    # Include product types for dropdown
            })

        if 'success' in result:
            messages.success(request, "Behavioral pattern saved successfully!")
            return redirect('behavioral_patterns_list')

    # Prepopulate the form with Time Bucket Entries and Product Types
    return render(request, 'ALM_APP/behavioral/create_behavioral_pattern.html', {
        'bucket_entries': bucket_entries,  # Pass time bucket entries for prepopulating
        'product_types': product_types,    # Pass product types for dropdown
    })


# View for Behavioral Pattern List

@login_required
def behavioral_patterns_list(request):
    patterns = BehavioralPatternConfig.objects.all().order_by(
        '-created_at')  # Fetching all patterns sorted by newest first
    return render(request, 'ALM_APP/behavioral/behavioral_patterns_list.html', {'patterns': patterns})

# View for Editing Behavioral Pattern
# In your views.py

@login_required
def edit_behavioral_pattern(request, id):
    try:
        # Get the pattern to edit
        pattern = BehavioralPatternConfig.objects.get(id=id)

        # Fetch product types from Ldn_Product_Master for the dropdown
        product_types = Ldn_Product_Master.objects.values_list('v_prod_type', flat=True).distinct()

        if request.method == 'POST':
            # Extract data from POST request
            tenors = request.POST.getlist('tenor[]')
            multipliers = request.POST.getlist('multiplier[]')
            percentages = request.POST.getlist('percentage[]')

            # Validate data lengths
            if not (len(tenors) == len(multipliers) == len(percentages)):
                return render(request, 'ALM_APP/behavioral/edit_behavioral_pattern.html', {
                    'error': "Mismatch in the number of entries for Tenors, Multipliers, and Percentages.",
                    'v_prod_type': request.POST.get('v_prod_type', pattern.v_prod_type),
                    'description': request.POST.get('description', pattern.description),
                    'tenors': tenors,
                    'multipliers': multipliers,
                    'percentages': percentages,
                    'product_types': product_types,  # Pass product types for the dropdown
                })

            # Call the utility function to update the pattern
            result = update_behavioral_pattern_from_form_data(request, pattern)

            if 'error' in result:
                return render(request, 'ALM_APP/behavioral/edit_behavioral_pattern.html', {
                    'error': result['error'],
                    'v_prod_type': request.POST.get('v_prod_type', pattern.v_prod_type),
                    'description': request.POST.get('description', pattern.description),
                    'tenors': tenors,
                    'multipliers': multipliers,
                    'percentages': percentages,
                    'product_types': product_types,  # Pass product types for the dropdown
                })

            # If successful, display the success message and redirect
            messages.success(request, "Behavioral pattern updated successfully!")
            return redirect('behavioral_patterns_list')

        # If GET request, prepopulate the form with the current data
        return render(request, 'ALM_APP/behavioral/edit_behavioral_pattern.html', {
            'v_prod_type': pattern.v_prod_type,
            'description': pattern.description,
            'tenors': [entry.tenor for entry in pattern.entries.all()],
            'multipliers': [entry.multiplier for entry in pattern.entries.all()],
            'percentages': [entry.percentage for entry in pattern.entries.all()],
            'product_types': product_types,  # Pass product types for the dropdown
        })

    except BehavioralPatternConfig.DoesNotExist:
        messages.error(request, "Behavioral pattern not found.")
        return redirect('behavioral_patterns_list')

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('behavioral_patterns_list')


# View for Deleting Behavioral Pattern
@login_required
def delete_behavioral_pattern(request, id):
    if request.method == 'POST':
        result = delete_behavioral_pattern_by_id(id)

        if 'error' in result:
            messages.error(request, result['error'])
        else:
            messages.success(
                request, "Behavioral pattern deleted successfully!")

        return redirect('behavioral_patterns_list')

@login_required
def view_behavioral_pattern(request, id):
    behavioral_pattern = get_object_or_404(BehavioralPatternConfig, id=id)
    pattern_entries = BehavioralPatternEntry.objects.filter(
        pattern=behavioral_pattern).order_by('order')

    return render(request, 'ALM_APP/behavioral/view_behavioral_pattern.html', {
        'behavioral_pattern': behavioral_pattern,
        'pattern_entries': pattern_entries
    })


@login_required
def create_time_bucket(request):
    # Check if a Time Bucket Definition already exists
    if TimeBucketDefinition.objects.exists():
        messages.error(request, "Only one Time Bucket Definition is allowed.")
        # Redirect to the list page if one already exists
        return redirect('time_bucket_list')

    if request.method == 'POST':
        # Call the function to process the form data
        result = define_time_bucket_from_form_data(request)

        # Check if there is an error
        if 'error' in result:
            # Use Django messages framework to display the error on the frontend
            messages.error(request, result['error'])

            # Return the form with existing data to repopulate the form fields
            return render(request, 'ALM_APP/time_buckets/create_time_bucket.html', {
                'name': request.POST.get('name'),
                'frequencies': request.POST.getlist('frequency[]'),
                'multipliers': request.POST.getlist('multiplier[]'),
                'start_dates': request.POST.getlist('start_date[]'),
                'end_dates': request.POST.getlist('end_date[]')
            })

        # If no error, assume success and redirect using the PRG pattern
        if 'success' in result:
            messages.success(request, "Time bucket saved successfully!")
            # Redirect to a success page or list
            return redirect('time_bucket_list')

    # If it's not a POST request, render the form
    # Redirect to the time buckets list page
    return render(request, 'ALM_APP/time_buckets/create_time_bucket.html')


# View for Listing Time Buckets
@login_required
def time_buckets_list(request):
    time_buckets = TimeBucketDefinition.objects.all().order_by(
        '-created_at')  # Fetching all time buckets sorted by newest first
    return render(request, 'ALM_APP/time_buckets/time_bucket_list.html', {'time_buckets': time_buckets})


# View for Editing a Time Bucket
@login_required
def edit_time_bucket(request, id):
    try:
        # Get the time bucket definition to edit
        time_bucket = TimeBucketDefinition.objects.get(id=id)

        if request.method == 'POST':
            # Call the utility function to update the time bucket
            result = update_time_bucket_from_form_data(request, time_bucket)

            if 'error' in result:
                return render(request, 'ALM_APP/time_buckets/edit_time_bucket.html', {
                    'error': result['error'],
                    'name': time_bucket.name,
                    'frequencies': [entry.frequency for entry in time_bucket.buckets.all()],
                    'multipliers': [entry.multiplier for entry in time_bucket.buckets.all()],
                    'start_dates': [entry.start_date for entry in time_bucket.buckets.all()],
                    'end_dates': [entry.end_date for entry in time_bucket.buckets.all()],
                })

            # If successful, display the success message and redirect
            messages.success(request, "Time bucket updated successfully!")
            # Redirect back to the time buckets list
            return redirect('time_bucket_list')

        # If GET request, prepopulate the form with the current data
        return render(request, 'ALM_APP/time_buckets/edit_time_bucket.html', {
            'name': time_bucket.name,
            'frequencies': [entry.frequency for entry in time_bucket.buckets.all()],
            'multipliers': [entry.multiplier for entry in time_bucket.buckets.all()],
            'start_dates': [entry.start_date for entry in time_bucket.buckets.all()],
            'end_dates': [entry.end_date for entry in time_bucket.buckets.all()],
        })

    except TimeBucketDefinition.DoesNotExist:
        messages.error(request, "Time bucket not found.")
        return redirect('time_bucket_list')


# View for Deleting a Time Bucket
@login_required
def delete_time_bucket(request, id):
    if request.method == 'POST':
        result = delete_time_bucket_by_id(id)

        if 'error' in result:
            messages.error(request, result['error'])
        else:
            messages.success(request, "Time bucket deleted successfully!")

        return redirect('time_bucket_list')


@login_required
def view_time_bucket(request, id):
    # Retrieve the specific Time Bucket Definition
    time_bucket = get_object_or_404(TimeBucketDefinition, id=id)

    # Pass the time bucket and its entries to the template
    return render(request, 'ALM_APP/time_buckets/view_time_bucket.html', {
        'time_bucket': time_bucket,
        'buckets': time_bucket.buckets.all().order_by('serial_number')
    })


# ProductFilter Views
class ProductFilterListView(ListView):
    model = ProductFilter
    template_name = 'ALM_APP/filters/filter_list.html'
    context_object_name = 'filters'


class ProductFilterCreateView(CreateView):
    model = ProductFilter
    form_class = ProductFilterForm
    template_name = 'ALM_APP/filters/filter_form.html'
    success_url = reverse_lazy('product_filter_list')

    def form_valid(self, form):
        create_or_update_filter(data=form.cleaned_data)
        messages.success(self.request, 'Product filter created successfully.')
        return redirect(self.success_url)


class ProductFilterUpdateView(UpdateView):
    model = ProductFilter
    form_class = ProductFilterForm
    template_name = 'ALM_APP/filters/filter_update.html'
    success_url = reverse_lazy('product_filter_list')

    def form_valid(self, form):
        create_or_update_filter(
            filter_id=self.object.id, data=form.cleaned_data)
        messages.success(self.request, 'Product filter updated successfully.')
        return redirect(self.success_url)


class ProductFilterDeleteView(View):
    success_url = reverse_lazy('product_filter_list')

    def post(self, request, *args, **kwargs):
        # Retrieve the filter to delete
        product_filter = get_object_or_404(ProductFilter, id=kwargs['pk'])
        delete_filter(filter_id=product_filter.id)
        messages.success(request, 'Product filter deleted successfully.')
        return redirect(self.success_url)


class ProductFilterDetailView(DetailView):
    model = ProductFilter
    template_name = 'ALM_APP/filters/filter_detail.html'
    context_object_name = 'filter'

    def get_object(self):
        # Fetch the filter based on ID or raise a 404 error
        filter_id = self.kwargs.get('pk')
        return get_object_or_404(ProductFilter, pk=filter_id)


# Process Views
class ProcessListView(ListView):
    model = Process
    template_name = 'ALM_APP/filters/process_list.html'
    context_object_name = 'processes'



@login_required
# def process_create_view(request):
#     step = request.session.get('step', 1)
#     print(f"\n=== Current Step: {step} ===")

#     # Step 1: Define Process Name and Description
#     if step == 1:
#         if request.method == 'POST':
#             process_name = request.POST.get('process_name')
#             process_description = request.POST.get('description')
#             use_behavioral_patterns = request.POST.get(
#                 'use_behavioral_patterns')

#             # Validate Process Name
#             if not process_name:
#                 messages.error(request, "Process name is required.")
#                 return render(request, 'ALM_APP/filters/process_create.html', {'step': step})

#             # Save details in session and proceed
#             request.session['process_name'] = process_name
#             request.session['process_description'] = process_description
#             request.session['use_behavioral_patterns'] = use_behavioral_patterns

#             # If behavioral patterns are used, skip to Step 3
#             if use_behavioral_patterns == 'yes':
#                 request.session['step'] = 3
#             else:
#                 request.session['step'] = 2  # Proceed to filter selection
#             return redirect('process_create')

#         return render(request, 'ALM_APP/filters/process_create.html', {'step': step})

#     # Step 2: Select Filters for the Process
#     elif step == 2:
#         filters = ProductFilter.objects.all()
#         if request.method == 'POST':
#             if 'previous' in request.POST:
#                 request.session['step'] = 1
#                 return redirect('process_create')
#             else:
#                 selected_filters = request.POST.getlist('filters')
#                 request.session['selected_filters'] = selected_filters
#                 request.session['step'] = 3
#                 return redirect('process_create')

#         return render(request, 'ALM_APP/filters/process_create.html', {'step': step, 'filters': filters})

#     # Step 3: Confirm and Execute
#     elif step == 3:
#         process_name = request.session.get('process_name')
#         process_description = request.session.get('process_description')
#         use_behavioral_patterns = request.session.get(
#             'use_behavioral_patterns')
#         selected_filters = request.session.get('selected_filters', [])
#         filters = ProductFilter.objects.filter(id__in=selected_filters)

#         if request.method == 'POST':
#             if 'previous' in request.POST:
#                 request.session['step'] = 2 if use_behavioral_patterns == 'no' else 1
#                 return redirect('process_create')
#             else:
#                 # Save the process
#                 process = finalize_process_creation(request)
#                 messages.success(request, f"Process '{
#                                  process.process_name}' created successfully.")

#                 # Clear session
#                 request.session.pop('process_name', None)
#                 request.session.pop('process_description', None)
#                 request.session.pop('use_behavioral_patterns', None)
#                 request.session.pop('selected_filters', None)
#                 request.session.pop('step', None)

#                 return redirect('processes_list')

#         return render(request, 'ALM_APP/filters/process_create.html', {
#             'step': step,
#             'process_name': process_name,
#             'process_description': process_description,
#             'selected_filters': filters,
#             'use_behavioral_patterns': use_behavioral_patterns
#         })

#     request.session['step'] = 1
#     return redirect('process_create')


def process_create_view(request):
    step = request.session.get('step', 1)
    print(f"\n=== Current Step: {step} ===")

    # Step 1: Define Process Name and Description
    if step == 1:
        if request.method == 'POST':
            process_name = request.POST.get('process_name')
            process_description = request.POST.get('description')
            use_behavioral_patterns = request.POST.get(
                'use_behavioral_patterns')

            # Validate Process Name
            if not process_name:
                messages.error(request, "Process name is required.")
                return render(request, 'ALM_APP/filters/process_create.html', {'step': step})

            # Save details in session and proceed
            request.session['process_name'] = process_name
            request.session['process_description'] = process_description
            request.session['use_behavioral_patterns'] = use_behavioral_patterns

            # If behavioral patterns are used, skip to Step 3
            if use_behavioral_patterns == 'yes':
                request.session['step'] = 3
            else:
                request.session['step'] = 2  # Proceed to filter selection
            return redirect('process_create')

        return render(request, 'ALM_APP/filters/process_create.html', {'step': step})

    # Step 2: Select Filters for the Process
    elif step == 2:
        filters = ProductFilter.objects.all()
        if request.method == 'POST':
            if 'previous' in request.POST:
                request.session['step'] = 1
                return redirect('process_create')
            else:
                selected_filters = request.POST.getlist('filters')
                request.session['selected_filters'] = selected_filters
                request.session['step'] = 3
                return redirect('process_create')

        return render(request, 'ALM_APP/filters/process_create.html', {'step': step, 'filters': filters})

    # Step 3: Confirm and Execute
    elif step == 3:
        process_name = request.session.get('process_name')
        process_description = request.session.get('process_description')
        use_behavioral_patterns = request.session.get(
            'use_behavioral_patterns')
        selected_filters = request.session.get('selected_filters', [])
        filters = ProductFilter.objects.filter(id__in=selected_filters)

        if request.method == 'POST':
            if 'previous' in request.POST:
                request.session['step'] = 2 if use_behavioral_patterns == 'no' else 1
                return redirect('process_create')
            else:
                # Save the process
                process = finalize_process_creation(request)
                messages.success(request, f"Process '{
                                 process.process_name}' created successfully.")

                # Clear session
                request.session.pop('process_name', None)
                request.session.pop('process_description', None)
                request.session.pop('use_behavioral_patterns', None)
                request.session.pop('selected_filters', None)
                request.session.pop('step', None)

                return redirect('processes_list')

        return render(request, 'ALM_APP/filters/process_create.html', {
            'step': step,
            'process_name': process_name,
            'process_description': process_description,
            'selected_filters': filters,
            'use_behavioral_patterns': use_behavioral_patterns
        })

    request.session['step'] = 1
    return redirect('process_create')


##################################################################################################################

@login_required
def execute_alm_process_view(request):
    if request.method == 'POST':
        process_id = request.POST.get('process_id')
        fic_mis_date = request.POST.get('fic_mis_date')

        try:
            datetime.strptime(fic_mis_date, "%Y-%m-%d")
        except ValueError:
            messages.error(
                request, "Invalid date format. Please use YYYY-MM-DD.")
            return redirect('processes_list')

        process = get_object_or_404(Process, id=process_id)

        try:
            if process.uses_behavioral_patterns:
                # If the process uses behavioral patterns, skip filtering and use the behavioral patterns directly
                calculate_behavioral_pattern_distribution(
                    process.name, fic_mis_date)
            else:
                # Normal process without behavioral patterns
                calculate_time_buckets_and_spread(process.name, fic_mis_date)

            messages.success(request, f"Process '{
                             process.name}' executed successfully.")
        except Exception as e:
            messages.error(request, f"Error executing process: {e}")

        return redirect('processes_list')

    return render(request, 'ALM_APP/filters/process_execute.html')



##################################################################################################################################

def ProcessUpdateView(request, process_id):
    process = get_object_or_404(Process, id=process_id)
    step = request.session.get('edit_step', 1)
    print(f"\n=== Current Edit Step: {step} ===")

    if step == 1:
        if request.method == 'POST':
            process_name = request.POST.get('name')
            process_description = request.POST.get('description')
            use_behavioral_patterns = request.POST.get('use_behavioral_patterns')

            if not process_name:
                messages.error(request, "Process name is required.")
                return render(request, 'ALM_APP/filters/process_edit.html', {
                    'step': step,
                    'process': process,
                })

            request.session['edit_process_name'] = process_name
            request.session['edit_process_description'] = process_description
            request.session['edit_use_behavioral_patterns'] = use_behavioral_patterns

            if use_behavioral_patterns == 'yes':
                request.session['edit_step'] = 3
            else:
                request.session['edit_step'] = 2
            return redirect('process_update', process_id=process_id)

        return render(request, 'ALM_APP/filters/process_edit.html', {
            'step': step,
            'process': process,
            'process_name': process.name,
            'process_description': process.description,
            'use_behavioral_patterns': 'yes' if process.uses_behavioral_patterns else 'no',
        })

    elif step == 2:
        filters = ProductFilter.objects.all()
        selected_filters = request.session.get('edit_selected_filters', process.filters.values_list('id', flat=True))

        if request.method == 'POST':
            if 'previous' in request.POST:
                request.session['edit_step'] = 1
                return redirect('process_update', process_id=process_id)
            else:
                selected_filters = request.POST.getlist('filters')
                request.session['edit_selected_filters'] = selected_filters
                request.session['edit_step'] = 3
                return redirect('process_update', process_id=process_id)

        return render(request, 'ALM_APP/filters/process_edit.html', {
            'step': step,
            'process': process,
            'filters': filters,
            'selected_filters': selected_filters,
        })

    elif step == 3:
        process_name = request.session.get('edit_process_name', process.name)
        process_description = request.session.get('edit_process_description', process.description)
        use_behavioral_patterns = request.session.get('edit_use_behavioral_patterns', 'no')
        selected_filters = request.session.get('edit_selected_filters', process.filters.values_list('id', flat=True))
        filters = ProductFilter.objects.filter(id__in=selected_filters)

        if request.method == 'POST':
            if 'previous' in request.POST:
                request.session['edit_step'] = 2 if use_behavioral_patterns == 'no' else 1
                return redirect('process_update', process_id=process_id)
            else:
                finalize_process_update(process, {
                    'name': process_name,
                    'description': process_description,
                    'use_behavioral_patterns': use_behavioral_patterns,
                    'filters': filters,
                })
                messages.success(request, f"Process '{process_name}' updated successfully.")

                # Clear session
                request.session.pop('edit_process_name', None)
                request.session.pop('edit_process_description', None)
                request.session.pop('edit_use_behavioral_patterns', None)
                request.session.pop('edit_selected_filters', None)
                request.session.pop('edit_step', None)

                return redirect('processes_list')

        return render(request, 'ALM_APP/filters/process_edit.html', {
            'step': step,
            'process': process,
            'process_name': process_name,
            'process_description': process_description,
            'selected_filters': filters,
            'use_behavioral_patterns': use_behavioral_patterns,
        })

    request.session['edit_step'] = 1
    return redirect('process_update', process_id=process_id)


#########################################################################
def processes_view(request, process_id):
    """
    Displays the details of a specific process.
    """
    process = get_object_or_404(Process, id=process_id)
    filters = process.filters.all()  # Assuming Process has a filters relationship
    return render(request, 'ALM_APP/filters/process_detail.html', {
        'process': process,
        'filters': filters,
    })



####################################################################################

def ProcessDeleteView(request, process_id):
    if request.method == 'POST':
        process = get_object_or_404(Process, id=process_id)
        try:
            process.delete()
            messages.success(request, 'Process deleted successfully.')
        except Exception as e:
            messages.error(request, f'Failed to delete process: {e}')
        return redirect('processes_list')

    elif request.method == 'GET':
        process = get_object_or_404(Process, id=process_id)
        return render(request, 'ALM_APP/filters/process_confirm_delete.html', {'object': process})

    return HttpResponseForbidden("Invalid request method")

















################################################################################################################




@login_required

def liquidity_gap_report_base(request):
    if request.GET:
        session_filters = request.session.get('filters_base', {})
        new_params = request.GET.dict()
        merged_filters = {**session_filters, **new_params}
        filters = merged_filters
        request.session['filters_base'] = filters
    else:
        filters = request.session.get('filters_base', {})

    form = LiquidityGapReportFilterForm(filters or None)
    base_queryset = LiquidityGapResultsBase.objects.all()

    # Fetch fic_mis_date from the form or fallback to the latest date
    fic_mis_date = form.cleaned_data.get('fic_mis_date') if form.is_valid() else None
    if not fic_mis_date:
        fic_mis_date = get_latest_fic_mis_date()
        if not fic_mis_date:
            messages.error(request, "No data available for the selected filters.")
            return render(
                request,
                'ALM_APP/reports/liquidity_gap_report_base.html',
                {'form': form}
            )

    # Get date buckets for fic_mis_date
    date_buckets = get_date_buckets(fic_mis_date)
    if not date_buckets.exists():
        messages.error(request, "No date buckets available for the selected filters.")
        return render(
            request,
            'ALM_APP/reports/liquidity_gap_report_base.html',
            {'form': form}
        )

    # Read drill-down params
    drill_down_product = request.GET.get('drill_down_product')
    drill_down_splits = request.GET.get('drill_down_splits')

    # Read selected_currency to lock page to a specific currency
    selected_currency = request.GET.get('selected_currency', None)

    # Filter the base queryset by form + fic_mis_date
    base_queryset = filter_queryset_by_form(form, base_queryset).filter(fic_mis_date=fic_mis_date)

    # Prepare a data structure for each currency
    currency_data = defaultdict(lambda: {
        'inflow_data': {},
        'outflow_data': {},
        'net_liquidity_gap': {},
        'net_gap_percentage': {},
        'cumulative_gap': {},
        'first_inflow_product': None,
        'remaining_inflow_data': {},
        'first_outflow_product': None,
        'remaining_outflow_data': {},
        'aggregated_product_details': None,
        'aggregated_split_details': None,
    })

    # Determine which currencies to iterate over
    if selected_currency:
        currencies = [selected_currency]
    else:
        currencies = base_queryset.values_list('v_ccy_code', flat=True).distinct()

    # Build currency-specific data
    for currency in currencies:
        currency_queryset = base_queryset.filter(v_ccy_code=currency)

        # Further localize the drill-down to this currency
        local_drill_down_qs = currency_queryset
        if drill_down_product:
            local_drill_down_qs = local_drill_down_qs.filter(v_prod_type=drill_down_product)
        if drill_down_splits:
            local_drill_down_qs = local_drill_down_qs.filter(v_product_name=drill_down_splits)

        # ----------------------------
        # AGGREGATED DRILL-DOWN LOGIC
        # ----------------------------
        local_aggregated_product_details = None
        local_aggregated_split_details = None

        if drill_down_splits:
            drill_down_splits_details = list(
                local_drill_down_qs
                .values('v_product_splits', 'bucket_number')
                .annotate(
                    inflows_total=Sum('inflows'),
                    outflows_total=Sum('outflows'),
                    total=Sum(F('inflows') - F('outflows'))
                )
            )
            grouped_split_data = defaultdict(lambda: {b['bucket_number']: 0 for b in date_buckets})
            for detail in drill_down_splits_details:
                split_name = detail.get('v_product_splits')
                bucket_number = detail['bucket_number']
                grouped_split_data[split_name][bucket_number] += detail.get('total', 0)

            local_aggregated_split_details = [
                {
                    'v_product_splits': split_name,
                    'buckets': buckets,
                    'total': sum(buckets.values())
                }
                for split_name, buckets in grouped_split_data.items()
            ]

        elif drill_down_product:
            drill_down_details = list(
                local_drill_down_qs
                .values('v_product_name', 'bucket_number')
                .annotate(
                    inflows_total=Sum('inflows'),
                    outflows_total=Sum('outflows'),
                    total=Sum(F('inflows') - F('outflows'))
                )
            )
            grouped_data = defaultdict(lambda: {b['bucket_number']: 0 for b in date_buckets})
            for detail in drill_down_details:
                product_name = detail.get('v_product_name')
                bucket_number = detail['bucket_number']
                grouped_data[product_name][bucket_number] += detail.get('total', 0)

            local_aggregated_product_details = [
                {
                    'v_product_name': product_name,
                    'buckets': buckets,
                    'total': sum(buckets.values())
                }
                for product_name, buckets in grouped_data.items()
            ]

        inflow_data, outflow_data = prepare_inflow_outflow_data(local_drill_down_qs)
        net_liquidity_gap, net_gap_percentage, cumulative_gap = calculate_totals(
            date_buckets, inflow_data, outflow_data
        )

        for product, buckets in inflow_data.items():
            inflow_data[product]['total'] = sum(
                buckets.get(b['bucket_number'], 0) for b in date_buckets
            )
        for product, buckets in outflow_data.items():
            outflow_data[product]['total'] = sum(
                buckets.get(b['bucket_number'], 0) for b in date_buckets
            )

        net_liquidity_gap['total'] = sum(
            net_liquidity_gap.get(b['bucket_number'], 0) for b in date_buckets
        )
        net_gap_percentage['total'] = (
            sum(net_gap_percentage.get(b['bucket_number'], 0) for b in date_buckets) / len(date_buckets)
        ) if len(date_buckets) > 0 else 0

        last_bucket = date_buckets.last()
        cumulative_gap['total'] = (
            cumulative_gap.get(last_bucket['bucket_number'], 0) if last_bucket else 0
        )

        if inflow_data:
            first_inflow_product = list(inflow_data.items())[0]
            remaining_inflow_data = inflow_data.copy()
            remaining_inflow_data.pop(first_inflow_product[0], None)
        else:
            first_inflow_product = None
            remaining_inflow_data = {}

        if outflow_data:
            first_outflow_product = list(outflow_data.items())[0]
            remaining_outflow_data = outflow_data.copy()
            remaining_outflow_data.pop(first_outflow_product[0], None)
        else:
            first_outflow_product = None
            remaining_outflow_data = {}

        currency_data[currency].update({
            'inflow_data': inflow_data,
            'outflow_data': outflow_data,
            'first_inflow_product': first_inflow_product,
            'remaining_inflow_data': remaining_inflow_data,
            'first_outflow_product': first_outflow_product,
            'remaining_outflow_data': remaining_outflow_data,
            'net_liquidity_gap': net_liquidity_gap,
            'net_gap_percentage': net_gap_percentage,
            'cumulative_gap': cumulative_gap,
            'aggregated_product_details': local_aggregated_product_details,
            'aggregated_split_details': local_aggregated_split_details,
        })

    context = {
        'form': form,
        'fic_mis_date': fic_mis_date,
        'date_buckets': date_buckets,
        'currency_data': dict(currency_data),
        'total_columns': len(date_buckets) + 3,
        'drill_down_product': drill_down_product,
        'drill_down_splits': drill_down_splits,
        'selected_currency': selected_currency,
    }

    return render(request, 'ALM_APP/reports/liquidity_gap_report_base.html', context)






########################################################################################################################
def liquidity_gap_report_cons(request):
    if request.GET:
        session_filters = request.session.get('filters_cons', {})
        new_params = request.GET.dict()
        merged_filters = {**session_filters, **new_params}
        filters = merged_filters
        request.session['filters_cons'] = filters
    else:
        filters = request.session.get('filters_cons', {})
    form = LiquidityGapReportFilterForm_cons(filters or None)
    cons_queryset = LiquidityGapResultsCons.objects.all()

    fic_mis_date = form.cleaned_data.get('fic_mis_date') if form.is_valid() else None
    if not fic_mis_date:
        fic_mis_date = get_latest_fic_mis_date()
        if not fic_mis_date:
            messages.error(request, "No data available for the selected filters.")
            return render(request, 'ALM_APP/reports/liquidity_gap_report_cons.html', {'form': form})

    date_buckets = get_date_buckets(fic_mis_date)
    if not date_buckets.exists():
        messages.error(request, "No date buckets available for the selected filters.")
        return render(request, 'ALM_APP/reports/liquidity_gap_report_cons.html', {'form': form})

    drill_down_product_cons = request.GET.get('drill_down_product_cons', None)
    drill_down_splits_cons = request.GET.get('drill_down_splits_cons', None)

    cons_queryset = filter_queryset_by_form(form, cons_queryset).filter(fic_mis_date=fic_mis_date)

    drill_down_details_cons = None
    drill_down_splits_details_cons = None
    aggregated_product_details_cons = None
    aggregated_split_details_cons = None

    if drill_down_splits_cons:
        drill_down_splits_details_cons = list(
            cons_queryset.filter(v_product_name=drill_down_splits_cons)
            .values('v_product_splits', 'bucket_number')
            .annotate(
                inflows_total=Sum('inflows'),
                outflows_total=Sum('outflows'),
                total=Sum(F('inflows') - F('outflows'))
            )
        )
        grouped_split_data_cons = defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})
        for detail in drill_down_splits_details_cons:
            split_name = detail.get('v_product_splits')
            bucket_number = detail['bucket_number']
            grouped_split_data_cons[split_name][bucket_number] += detail.get('total', 0)

        aggregated_split_details_cons = [
            {
                'v_product_splits': split_name,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for split_name, buckets in grouped_split_data_cons.items()
        ]

    elif drill_down_product_cons:
        drill_down_details_cons = list(
            cons_queryset.filter(v_prod_type=drill_down_product_cons)
            .values('v_product_name', 'bucket_number')
            .annotate(
                inflows_total=Sum('inflows'),
                outflows_total=Sum('outflows'),
                total=Sum(F('inflows') - F('outflows'))
            )
        )
        grouped_data_cons = defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})
        for detail in drill_down_details_cons:
            product_name = detail.get('v_product_name')
            bucket_number = detail['bucket_number']
            grouped_data_cons[product_name][bucket_number] += detail.get('total', 0)

        aggregated_product_details_cons = [
            {
                'v_product_name': product_name,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for product_name, buckets in grouped_data_cons.items()
        ]

    cons_inflow_data, cons_outflow_data = prepare_inflow_outflow_data(cons_queryset)
    cons_net_liquidity_gap, cons_net_gap_percentage, cons_cumulative_gap = calculate_totals(
        date_buckets, cons_inflow_data, cons_outflow_data
    )

    if cons_inflow_data:
        cons_first_inflow_product = list(cons_inflow_data.items())[0]
        cons_remaining_inflow_data = cons_inflow_data.copy()
        cons_remaining_inflow_data.pop(cons_first_inflow_product[0], None)
    else:
        cons_first_inflow_product = None
        cons_remaining_inflow_data = {}

    if cons_outflow_data:
        cons_first_outflow_product = list(cons_outflow_data.items())[0]
        cons_remaining_outflow_data = cons_outflow_data.copy()
        cons_remaining_outflow_data.pop(cons_first_outflow_product[0], None)
    else:
        cons_first_outflow_product = None
        cons_remaining_outflow_data = {}

    cons_data = {
        'inflow_data': cons_inflow_data,
        'outflow_data': cons_outflow_data,
        'first_inflow_product': cons_first_inflow_product,
        'remaining_inflow_data': cons_remaining_inflow_data,
        'first_outflow_product': cons_first_outflow_product,
        'remaining_outflow_data': cons_remaining_outflow_data,
        'net_liquidity_gap': cons_net_liquidity_gap,
        'net_gap_percentage': cons_net_gap_percentage,
        'cumulative_gap': cons_cumulative_gap,
    }

    print("Consolidated Data:", cons_data)

    drill_cons_data = None
    if drill_down_product_cons or drill_down_splits_cons:
        if drill_down_product_cons:
            drill_queryset = cons_queryset.filter(v_prod_type=drill_down_product_cons)
        elif drill_down_splits_cons:
            drill_queryset = cons_queryset.filter(v_product_name=drill_down_splits_cons)

        drill_inflow_data, drill_outflow_data = prepare_inflow_outflow_data(drill_queryset)
        drill_net_liquidity_gap, drill_net_gap_percentage, drill_cumulative_gap = calculate_totals(
            date_buckets, drill_inflow_data, drill_outflow_data
        )

        drill_cons_data = {
            'inflow_data': drill_inflow_data,
            'outflow_data': drill_outflow_data,
            'net_liquidity_gap': drill_net_liquidity_gap,
            'net_gap_percentage': drill_net_gap_percentage,
            'cumulative_gap': drill_cumulative_gap,
        }

    # Extract distinct currencies
    currency_data = cons_queryset.values_list('v_ccy_code', flat=True).distinct()


    context = {
        'form': form,
        'fic_mis_date': fic_mis_date,
        'date_buckets': date_buckets,
        'currency_data': currency_data,  # Pass currency data to the template
        'cons_data': cons_data,
        'drill_cons_data': drill_cons_data,
        'total_columns': len(date_buckets) + 3,
        'drill_down_product_cons': drill_down_product_cons,
        'drill_down_splits_cons': drill_down_splits_cons,
        'aggregated_product_details_cons': aggregated_product_details_cons,
        'aggregated_split_details_cons': aggregated_split_details_cons
    }

    return render(request, 'ALM_APP/reports/liquidity_gap_report_cons.html', context)




















# def liquidity_gap_report(request):
#     # Initialize the form with GET parameters
#     form = LiquidityGapReportFilterForm(request.GET or None)

#     # Start with the full queryset for base and consolidated results
#     base_queryset = LiquidityGapResultsBase.objects.all()
#     cons_queryset = LiquidityGapResultsCons.objects.all()

#     # Get fic_mis_date from the form or fallback to the latest date in Dim_Dates
#     fic_mis_date = form.cleaned_data.get(
#         'fic_mis_date') if form.is_valid() else None
#     if not fic_mis_date:
#         fic_mis_date = get_latest_fic_mis_date()
#         if not fic_mis_date:
#             messages.error(
#                 request, "No data available for the selected filters.")
#             return render(request, 'ALM_APP/reports/liquidity_gap_report.html', {'form': form})

#     # Get date buckets for fic_mis_date
#     date_buckets = get_date_buckets(fic_mis_date)
#     if not date_buckets.exists():
#         messages.error(
#             request, "No date buckets available for the selected filters.")
#         return render(request, 'ALM_APP/reports/liquidity_gap_report.html', {'form': form})

#     # Check if this is a drill-down request for product or splits
#     drill_down_product = request.GET.get('drill_down_product', None)
#     drill_down_splits = request.GET.get('drill_down_splits', None)
#     drill_down_product_cons = request.GET.get('drill_down_product_cons', None)
#     drill_down_splits_cons = request.GET.get('drill_down_splits_cons', None)

#     # Filter base and consolidated querysets by form fields
#     base_queryset = filter_queryset_by_form(
#         form, base_queryset).filter(fic_mis_date=fic_mis_date)
#     cons_queryset = filter_queryset_by_form(
#         form, cons_queryset).filter(fic_mis_date=fic_mis_date)

#     # Prepare drill-down details for products or splits on cons and base
#     drill_down_details = None
#     drill_down_splits_details = None
#     aggregated_product_details = None
#     aggregated_split_details = None

#     drill_down_details_cons = None
#     drill_down_splits_details_cons = None
#     aggregated_product_details_cons = None
#     aggregated_split_details_cons = None

#     if drill_down_splits:  # Drill-down for splits
#         drill_down_splits_details = list(
#             base_queryset.filter(v_product_name=drill_down_splits)
#             .values('v_product_splits', 'bucket_number')
#             .annotate(
#                 inflows_total=Sum('inflows'),
#                 outflows_total=Sum('outflows'),
#                 total=Sum(F('inflows') - F('outflows'))
#             )
#         )

#         # Group data by product splits and bucket
#         grouped_split_data = defaultdict(
#             lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})
#         for detail in drill_down_splits_details:
#             # Get product split name
#             split_name = detail.get('v_product_splits')
#             bucket_number = detail['bucket_number']  # Get bucket number
#             # Aggregate figures by bucket
#             grouped_split_data[split_name][bucket_number] += detail.get(
#                 'total', 0)

#         # Convert grouped data into a list for the template
#         aggregated_split_details = [
#             {
#                 'v_product_splits': split_name,  # Include split name
#                 'buckets': buckets,              # Figures grouped by buckets
#                 'total': sum(buckets.values())   # Total for the product split
#             }
#             for split_name, buckets in grouped_split_data.items()
#         ]

#     elif drill_down_product:  # Drill-down for product names
#         drill_down_details = list(
#             base_queryset.filter(v_prod_type=drill_down_product)
#             .values('v_product_name', 'bucket_number')
#             .annotate(
#                 inflows_total=Sum('inflows'),
#                 outflows_total=Sum('outflows'),
#                 total=Sum(F('inflows') - F('outflows'))
#             )
#         )

#         # Group data by product name and bucket
#         grouped_data = defaultdict(
#             lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})
#         for detail in drill_down_details:
#             product_name = detail.get('v_product_name')  # Get product name
#             bucket_number = detail['bucket_number']  # Get bucket number
#             # Aggregate figures by bucket
#             grouped_data[product_name][bucket_number] += detail.get('total', 0)

#         # Convert grouped data into a list for the template
#         aggregated_product_details = [
#             {
#                 'v_product_name': product_name,  # Include product name
#                 'buckets': buckets,              # Figures grouped by buckets
#                 'total': sum(buckets.values())   # Total for the product name
#             }
#             for product_name, buckets in grouped_data.items()
#         ]

#     if drill_down_splits_cons:  # Drill-down for splits in cons
#         drill_down_splits_details_cons = list(
#             cons_queryset.filter(v_product_name=drill_down_splits_cons)
#             .values('v_product_splits', 'bucket_number')
#             .annotate(
#                 inflows_total=Sum('inflows'),
#                 outflows_total=Sum('outflows'),
#                 total=Sum(F('inflows') - F('outflows'))
#             )
#         )

#         grouped_split_data_cons = defaultdict(
#             lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})
#         for detail in drill_down_splits_details_cons:
#             split_name = detail.get('v_product_splits')
#             bucket_number = detail['bucket_number']
#             grouped_split_data_cons[split_name][bucket_number] += detail.get(
#                 'total', 0)

#         aggregated_split_details_cons = [
#             {
#                 'v_product_splits': split_name,
#                 'buckets': buckets,
#                 'total': sum(buckets.values())
#             }
#             for split_name, buckets in grouped_split_data_cons.items()
#         ]

#     elif drill_down_product_cons:  # Drill-down for product names in cons
#         drill_down_details_cons = list(
#             cons_queryset.filter(v_prod_type=drill_down_product_cons)
#             .values('v_product_name', 'bucket_number')
#             .annotate(
#                 inflows_total=Sum('inflows'),
#                 outflows_total=Sum('outflows'),
#                 total=Sum(F('inflows') - F('outflows'))
#             )
#         )

#         grouped_data_cons = defaultdict(
#             lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})
#         for detail in drill_down_details_cons:
#             product_name = detail.get('v_product_name')
#             bucket_number = detail['bucket_number']
#             grouped_data_cons[product_name][bucket_number] += detail.get(
#                 'total', 0)

#         aggregated_product_details_cons = [
#             {
#                 'v_product_name': product_name,
#                 'buckets': buckets,
#                 'total': sum(buckets.values())
#             }
#             for product_name, buckets in grouped_data_cons.items()
#         ]

#     # Prepare base results
#     currency_data = defaultdict(lambda: {
#         'inflow_data': {}, 'outflow_data': {},
#         'net_liquidity_gap': {}, 'net_gap_percentage': {}, 'cumulative_gap': {},
#         'first_inflow_product': None, 'remaining_inflow_data': {},
#         'first_outflow_product': None, 'remaining_outflow_data': {}
#     })

#     currencies = base_queryset.values_list('v_ccy_code', flat=True).distinct()
#     for currency in currencies:
#         # Filter by currency
#         currency_queryset = base_queryset.filter(v_ccy_code=currency)

#         if drill_down_product:
#             currency_queryset = currency_queryset.filter(
#                 v_prod_type=drill_down_product)
#         if drill_down_splits:
#             currency_queryset = currency_queryset.filter(
#                 v_product_name=drill_down_splits)

#         inflow_data, outflow_data = prepare_inflow_outflow_data(
#             currency_queryset)

#         net_liquidity_gap, net_gap_percentage, cumulative_gap = calculate_totals(
#             date_buckets, inflow_data, outflow_data)

#         for product, buckets in inflow_data.items():
#             inflow_data[product]['total'] = sum(buckets.get(
#                 bucket['bucket_number'], 0) for bucket in date_buckets)

#         for product, buckets in outflow_data.items():
#             outflow_data[product]['total'] = sum(buckets.get(
#                 bucket['bucket_number'], 0) for bucket in date_buckets)

#         net_liquidity_gap['total'] = sum(net_liquidity_gap.get(
#             bucket['bucket_number'], 0) for bucket in date_buckets)
#         net_gap_percentage['total'] = (
#             sum(net_gap_percentage.get(bucket['bucket_number'], 0)
#                 for bucket in date_buckets) / len(date_buckets)
#         ) if len(date_buckets) > 0 else 0

#         last_bucket = date_buckets.last()
#         cumulative_gap['total'] = cumulative_gap.get(
#             last_bucket['bucket_number'], 0) if last_bucket else 0

#         if inflow_data:
#             first_inflow_product = list(inflow_data.items())[0]
#             remaining_inflow_data = inflow_data.copy()
#             remaining_inflow_data.pop(first_inflow_product[0], None)
#         else:
#             first_inflow_product = None
#             remaining_inflow_data = {}

#         if outflow_data:
#             first_outflow_product = list(outflow_data.items())[0]
#             remaining_outflow_data = outflow_data.copy()
#             remaining_outflow_data.pop(first_outflow_product[0], None)
#         else:
#             first_outflow_product = None
#             remaining_outflow_data = {}

#         currency_data[currency].update({
#             'inflow_data': inflow_data,
#             'outflow_data': outflow_data,
#             'first_inflow_product': first_inflow_product,
#             'remaining_inflow_data': remaining_inflow_data,
#             'first_outflow_product': first_outflow_product,
#             'remaining_outflow_data': remaining_outflow_data,
#             'net_liquidity_gap': net_liquidity_gap,
#             'net_gap_percentage': net_gap_percentage,
#             'cumulative_gap': cumulative_gap,
#         })

#     cons_inflow_data, cons_outflow_data = prepare_inflow_outflow_data(
#         cons_queryset)
#     cons_net_liquidity_gap, cons_net_gap_percentage, cons_cumulative_gap = calculate_totals(
#         date_buckets, cons_inflow_data, cons_outflow_data
#     )

#     if cons_inflow_data:
#         cons_first_inflow_product = list(cons_inflow_data.items())[0]
#         cons_remaining_inflow_data = cons_inflow_data.copy()
#         cons_remaining_inflow_data.pop(cons_first_inflow_product[0], None)
#     else:
#         cons_first_inflow_product = None
#         cons_remaining_inflow_data = {}

#     if cons_outflow_data:
#         cons_first_outflow_product = list(cons_outflow_data.items())[0]
#         cons_remaining_outflow_data = cons_outflow_data.copy()
#         cons_remaining_outflow_data.pop(cons_first_outflow_product[0], None)
#     else:
#         cons_first_outflow_product = None
#         cons_remaining_outflow_data = {}

#     cons_data = {
#         'inflow_data': cons_inflow_data,
#         'outflow_data': cons_outflow_data,
#         'first_inflow_product': cons_first_inflow_product,
#         'remaining_inflow_data': cons_remaining_inflow_data,
#         'first_outflow_product': cons_first_outflow_product,
#         'remaining_outflow_data': cons_remaining_outflow_data,
#         'net_liquidity_gap': cons_net_liquidity_gap,
#         'net_gap_percentage': cons_net_gap_percentage,
#         'cumulative_gap': cons_cumulative_gap,
#     }

#     print("Consolidated Data:", cons_data)

#     context = {
#         'form': form,
#         'fic_mis_date': fic_mis_date,
#         'date_buckets': date_buckets,
#         'currency_data': dict(currency_data),
#         'cons_data': cons_data,
#         'total_columns': len(date_buckets) + 3,
#         'drill_down_product': drill_down_product,
#         'drill_down_splits': drill_down_splits,
#         'drill_down_product_cons': drill_down_product_cons,
#         'drill_down_splits_cons': drill_down_splits_cons,
#         'aggregated_product_details': aggregated_product_details,
#         'aggregated_split_details': aggregated_split_details,
#         'aggregated_product_details_cons': aggregated_product_details_cons,
#         'aggregated_split_details_cons': aggregated_split_details_cons
#     }

#     return render(request, 'ALM_APP/reports/liquidity_gap_report.html', context)



##########################################################################################################################
@login_required

def export_liquidity_gap_to_excel(request):
    form = LiquidityGapReportFilterForm(request.GET or None)

    # Parse fic_mis_date from request
    raw_fic_mis_date = request.GET.get('fic_mis_date')
    fic_mis_date = parse_date(raw_fic_mis_date)

    if not fic_mis_date:
        return HttpResponseBadRequest(
            "Invalid date format. Supported formats: 'Aug. 31, 2024', 'August 31, 2024', "
            "'2024-08-31', '31-08-2024', '31/08/2024', '08/31/2024'."
        )

    # Base queryset filtered by fic_mis_date
    queryset = LiquidityGapResultsBase.objects.filter(fic_mis_date=fic_mis_date)
    queryset = filter_queryset_by_form(form, queryset)

    # Handle drill-down parameters
    drill_down_product = request.GET.get('drill_down_product')
    drill_down_splits = request.GET.get('drill_down_splits')

    if drill_down_product:
        queryset = queryset.filter(v_prod_type=drill_down_product)
    if drill_down_splits:
        queryset = queryset.filter(v_product_name=drill_down_splits)

    # Get date buckets
    date_buckets = get_date_buckets(fic_mis_date)
    if not date_buckets.exists():
        return HttpResponseBadRequest("No date buckets available for the selected date.")

    # Prepare headers based on drill-down level
    headers = ["Account Type", "Product"]
    if drill_down_splits:
        headers += ["v_product_name", "v_product_splits"]
    elif drill_down_product:
        headers += ["v_product_name"]

    headers += [
        f"{bucket['bucket_start_date'].strftime('%d-%b-%Y')} to {bucket['bucket_end_date'].strftime('%d-%b-%Y')}"
        for bucket in date_buckets
    ] + ["Total"]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"), right=Side(border_style="thin"),
        top=Side(border_style="thin"), bottom=Side(border_style="thin")
    )
    alignment_center = Alignment(horizontal="center", vertical="center")
    heading_font = Font(bold=True, size=14)

    # Overview scenario: separate sheets per currency
    if not drill_down_product and not drill_down_splits:
        # Group data by currency (v_ccy_code) and product type
        details = queryset.values('v_ccy_code', 'v_prod_type', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        currency_grouped = defaultdict(lambda: defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets}))
        for detail in details:
            currency = detail['v_ccy_code']  # Use v_ccy_code instead of currency
            prod_type = detail['v_prod_type']
            bucket_number = detail['bucket_number']
            currency_grouped[currency][prod_type][bucket_number] += detail.get('total', 0)

        # Initialize workbook and remove the default sheet
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        # Create a sheet for each currency
        for currency, products in currency_grouped.items():
            sheet = workbook.create_sheet(title=str(currency))
            total_columns = len(headers)

            # Dynamic heading for this currency sheet
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
            heading_cell = sheet.cell(row=1, column=1)
            heading_cell.value = f"Overview Report for {currency} on {fic_mis_date}"
            heading_cell.font = heading_font
            heading_cell.alignment = alignment_center

            # Write headers in row 2
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=2, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = border

            # Write aggregated data rows starting from row 3
            row_num = 3
            for prod_type, buckets in products.items():
                account_type = "Total Inflow" if sum(buckets.values()) >= 0 else "Total Outflow"
                row_prefix = [account_type, prod_type]
                row_data = row_prefix + [
                    buckets.get(bucket['bucket_number'], 0) for bucket in date_buckets
                ] + [sum(buckets.values())]

                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = alignment_center
                    cell.border = border
                row_num += 1

            # Compute summary rows for the current currency
            currency_qs = queryset.filter(v_ccy_code=currency)  # Filter by current currency
            inflow_data_cur, outflow_data_cur = prepare_inflow_outflow_data(currency_qs)
            net_liquidity_gap_cur, net_gap_percentage_cur, cumulative_gap_cur = calculate_totals(
                date_buckets, inflow_data_cur, outflow_data_cur
            )

            total_outflows_cur = sum(outflow_data_cur.get(bucket['bucket_number'], 0) for bucket in date_buckets)
            net_liquidity_gap_total_cur = sum(net_liquidity_gap_cur.get(bucket['bucket_number'], 0) for bucket in date_buckets)
            cumulative_gap_total_cur = sum(net_liquidity_gap_cur.get(bucket['bucket_number'], 0) for bucket in date_buckets)
            net_gap_percentage_total_cur = (net_liquidity_gap_total_cur / total_outflows_cur * 100) if total_outflows_cur else 0

            prefix_count = 2
            summary_rows = [
                ("Net Liquidity Gap", net_liquidity_gap_cur, net_liquidity_gap_total_cur),
                ("Net Gap as % of Total Outflows", net_gap_percentage_cur, net_gap_percentage_total_cur),
                ("Cumulative Gap", cumulative_gap_cur, cumulative_gap_total_cur),
            ]
            for label, summary_data, summary_total in summary_rows:
                row = [label] + [""] * (prefix_count - 1)
                row += [summary_data.get(bucket['bucket_number'], 0) for bucket in date_buckets]
                row.append(summary_total)

                for col_num, value in enumerate(row, 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.font = Font(bold=True)
                    cell.alignment = alignment_center
                    cell.border = border
                row_num += 1

            # Auto-adjust column widths for this sheet
            for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                max_length = 0
                for cell in col:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max_length + 2

        # Save and return the workbook for overview scenario
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="LiquidityGapReport_Overview_{fic_mis_date}.xlsx"'
        workbook.save(response)
        return response

    # Non-overview (drill-down) scenarios
    grouped_data = defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})

    if drill_down_splits:
        details = queryset.values('v_product_splits', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            split_name = detail['v_product_splits']
            bucket_number = detail['bucket_number']
            grouped_data[split_name][bucket_number] += detail.get('total', 0)
        aggregated_data = [
            {'identifier': split_name, 'buckets': buckets, 'total': sum(buckets.values())}
            for split_name, buckets in grouped_data.items()
        ]
    elif drill_down_product:
        details = queryset.values('v_product_name', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            product_name = detail['v_product_name']
            bucket_number = detail['bucket_number']
            grouped_data[product_name][bucket_number] += detail.get('total', 0)
        aggregated_data = [
            {'identifier': product_name, 'buckets': buckets, 'total': sum(buckets.values())}
            for product_name, buckets in grouped_data.items()
        ]
    else:
        details = queryset.values('v_prod_type', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            prod_type = detail['v_prod_type']
            bucket_number = detail['bucket_number']
            grouped_data[prod_type][bucket_number] += detail.get('total', 0)
        aggregated_data = [
            {'identifier': prod_type, 'buckets': buckets, 'total': sum(buckets.values())}
            for prod_type, buckets in grouped_data.items()
        ]

    # Calculate totals for summary rows for non-overview
    inflow_data, outflow_data = prepare_inflow_outflow_data(queryset)
    net_liquidity_gap, net_gap_percentage, cumulative_gap = calculate_totals(
        date_buckets, inflow_data, outflow_data
    )

    total_outflows = sum(outflow_data.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    net_liquidity_gap_total = sum(net_liquidity_gap.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    cumulative_gap_total = sum(cumulative_gap.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    net_gap_percentage_total = (net_liquidity_gap_total / total_outflows * 100) if total_outflows else 0

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Liquidity Gap Report"

    # Determine dynamic heading based on drill-down level
    if drill_down_splits:
        heading_text = f"Drill-Down Report for Split: {drill_down_splits} on {fic_mis_date}"
    elif drill_down_product:
        heading_text = f"Drill-Down Report for Product: {drill_down_product} on {fic_mis_date}"
    else:
        heading_text = f"Overview Report for {fic_mis_date}"

    total_columns = len(headers)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    heading_cell = sheet.cell(row=1, column=1)
    heading_cell.value = heading_text
    heading_cell.font = heading_font
    heading_cell.alignment = alignment_center

    # Write headers for non-overview
    sheet.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = border

    # Write aggregated data rows for non-overview
    row_num = 3
    for item in aggregated_data:
        if drill_down_splits:
            row_prefix = ["Total Flows", drill_down_product, drill_down_splits, item['identifier']]
        elif drill_down_product:
            row_prefix = ["Total Flows", drill_down_product, item['identifier']]
        else:
            account_type = "Total Inflow" if item['total'] >= 0 else "Total Outflow"
            row_prefix = [account_type, item['identifier']]

        row = row_prefix + [
            item['buckets'].get(bucket['bucket_number'], 0) for bucket in date_buckets
        ] + [item['total']]

        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = alignment_center
            cell.border = border
        row_num += 1

    if drill_down_splits:
        prefix_count = 4
    elif drill_down_product:
        prefix_count = 3
    else:
        prefix_count = 2

    summary_rows = [
        ("Net Liquidity Gap", net_liquidity_gap, net_liquidity_gap_total),
        ("Net Gap as % of Total Outflows", net_gap_percentage, net_gap_percentage_total),
        ("Cumulative Gap", cumulative_gap, cumulative_gap_total),
    ]

    for label, summary_data, summary_total in summary_rows:
        row = [label] + [""] * (prefix_count - 1)
        row += [summary_data.get(bucket['bucket_number'], 0) for bucket in date_buckets]
        row.append(summary_total)

        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(bold=True)
            cell.alignment = alignment_center
            cell.border = border
        row_num += 1

    for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_length = 0
        for cell in col:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="LiquidityGapReport_{fic_mis_date}.xlsx"'
    workbook.save(response)
    return response







########################################################################################################################################################################
def parse_date(date_str):
    """
    Parse the date string to handle multiple formats.
    """
    date_formats = [
        "%b. %d, %Y",    # 'Aug. 31, 2024'
        "%B %d, %Y",     # 'August 31, 2024'
        "%Y-%m-%d",      # '2024-08-31'
        "%d-%m-%Y",      # '31-08-2024'
        "%d/%m/%Y",      # '31/08/2024'
        "%m/%d/%Y"       # '08/31/2024'
    ]
    for date_format in date_formats:
        try:
            return datetime.strptime(date_str, date_format).date()
        except ValueError:
            continue
    return None




#################################################################################################################################


@login_required
def export_liquidity_gap_cons_to_excel(request):

    form = LiquidityGapReportFilterForm_cons(request.GET or None)

    # Parse fic_mis_date from request
    raw_fic_mis_date = request.GET.get('fic_mis_date')
    fic_mis_date = parse_date(raw_fic_mis_date)

    if not fic_mis_date:
        return HttpResponseBadRequest(
            "Invalid date format. Supported formats: 'Aug. 31, 2024', 'August 31, 2024', "
            "'2024-08-31', '31-08-2024', '31/08/2024', '08/31/2024'."
        )

    # Base queryset filtered by fic_mis_date
    queryset = LiquidityGapResultsCons.objects.filter(fic_mis_date=fic_mis_date)
    queryset = filter_queryset_by_form(form, queryset)

    # Handle drill-down parameters
    drill_down_product_cons = request.GET.get('drill_down_product_cons')
    drill_down_splits_cons = request.GET.get('drill_down_splits_cons')

    if drill_down_product_cons:
        queryset = queryset.filter(v_prod_type=drill_down_product_cons)
    if drill_down_splits_cons:
        queryset = queryset.filter(v_product_name=drill_down_splits_cons)

    # Get date buckets
    date_buckets = get_date_buckets(fic_mis_date)
    if not date_buckets.exists():
        return HttpResponseBadRequest("No date buckets available for the selected date.")

    # Prepare headers based on drill-down level
    headers = ["Account Type", "Product"]
    if drill_down_splits_cons:
        headers += ["v_product_name", "v_product_splits"]
    elif drill_down_product_cons:
        headers += ["v_product_name"]

    # Extend headers with bucket date ranges and Total column
    headers += [
        f"{bucket['bucket_start_date'].strftime('%d-%b-%Y')} to {bucket['bucket_end_date'].strftime('%d-%b-%Y')}"
        for bucket in date_buckets
    ] + ["Total"]

    # Prepare data for the current drill-down level
    grouped_data = defaultdict(lambda: {bucket['bucket_number']: 0 for bucket in date_buckets})

    if drill_down_splits_cons:
        details = queryset.values('v_product_splits', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            split_name = detail['v_product_splits']
            bucket_number = detail['bucket_number']
            grouped_data[split_name][bucket_number] += detail.get('total', 0)

        aggregated_data = [
            {
                'identifier': split_name,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for split_name, buckets in grouped_data.items()
        ]
    elif drill_down_product_cons:
        details = queryset.values('v_product_name', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            product_name = detail['v_product_name']
            bucket_number = detail['bucket_number']
            grouped_data[product_name][bucket_number] += detail.get('total', 0)

        aggregated_data = [
            {
                'identifier': product_name,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for product_name, buckets in grouped_data.items()
        ]
    else:
        # Overview level: Group by v_prod_type when no drill-down filters are applied
        details = queryset.values('v_prod_type', 'bucket_number').annotate(
            total=Sum(F('inflows') - F('outflows'))
        )
        for detail in details:
            prod_type = detail['v_prod_type']
            bucket_number = detail['bucket_number']
            grouped_data[prod_type][bucket_number] += detail.get('total', 0)

        aggregated_data = [
            {
                'identifier': prod_type,
                'buckets': buckets,
                'total': sum(buckets.values())
            }
            for prod_type, buckets in grouped_data.items()
        ]

    # Calculate totals for summary rows
    inflow_data, outflow_data = prepare_inflow_outflow_data(queryset)
    net_liquidity_gap, net_gap_percentage, cumulative_gap = calculate_totals(
        date_buckets, inflow_data, outflow_data
    )

    total_outflows = sum(outflow_data.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    net_liquidity_gap_total = sum(net_liquidity_gap.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    cumulative_gap_total = sum(net_liquidity_gap.get(bucket['bucket_number'], 0) for bucket in date_buckets)
    if total_outflows:
        net_gap_percentage_total = (net_liquidity_gap_total / total_outflows) * 100
    else:
        net_gap_percentage_total = 0

    # Initialize workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Liquidity Gap Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B5998", end_color="3B5998", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    alignment_center = Alignment(horizontal="center", vertical="center")
    heading_font = Font(bold=True, size=14)

    # Determine dynamic heading based on drill-down level
    if drill_down_splits_cons:
        heading_text = f"Drill-Down Report for Split: {drill_down_splits_cons} on {fic_mis_date}"
    elif drill_down_product_cons:
        heading_text = f"Drill-Down Report for Product: {drill_down_product_cons} on {fic_mis_date}"
    else:
        heading_text = f"Overview Report for {fic_mis_date}"

    total_columns = len(headers)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    heading_cell = sheet.cell(row=1, column=1)
    heading_cell.value = heading_text
    heading_cell.font = heading_font
    heading_cell.alignment = alignment_center

    # Write headers
    sheet.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = border

    # Write aggregated data rows
    row_num = 3
    for item in aggregated_data:
        if drill_down_splits_cons:
            row_prefix = ["Total Flows", drill_down_product_cons, drill_down_splits_cons, item['identifier']]
        elif drill_down_product_cons:
            row_prefix = ["Total Flows", drill_down_product_cons, item['identifier']]
        else:
            # For overview, determine account type based on total and use v_prod_type as product
            account_type = "Total Inflow" if item['total'] >= 0 else "Total Outflow"
            row_prefix = [account_type, item['identifier']]

        row = row_prefix + [
            item['buckets'].get(bucket['bucket_number'], 0) for bucket in date_buckets
        ] + [item['total']]

        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = alignment_center
            cell.border = border
        row_num += 1

    # Determine fixed prefix column count based on drill-down level
    if drill_down_splits_cons:
        prefix_count = 4
    elif drill_down_product_cons:
        prefix_count = 3
    else:
        prefix_count = 2

    summary_rows = [
        ("Net Liquidity Gap", net_liquidity_gap, net_liquidity_gap_total),
        ("Net Gap as % of Total Outflows", net_gap_percentage, net_gap_percentage_total),
        ("Cumulative Gap", cumulative_gap, cumulative_gap_total),
    ]

    # Write summary rows aligned with the rest of the data
    for label, summary_data, summary_total in summary_rows:
        row = [label] + [""] * (prefix_count - 1)
        row += [summary_data.get(bucket['bucket_number'], 0) for bucket in date_buckets]
        row.append(summary_total)

        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(bold=True)
            cell.alignment = alignment_center
            cell.border = border
        row_num += 1

    # Auto-adjust column widths
    for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_length = 0
        for cell in col:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="LiquidityGapReport_Cons_{fic_mis_date}.xlsx"'
    workbook.save(response)
    return response

















# from django.shortcuts import render
# from .models import LiquidityGapResultsBase, Dim_Dates
# from .forms import LiquidityGapReportFilterForm
# from django.db.models import Sum
# from django.contrib import messages  # Import for user messages
# from datetime import date

# def liquidity_gap_report(request):
#     # Initialize the form with GET parameters
#     form = LiquidityGapReportFilterForm(request.GET or None)

#     # Start with the full queryset
#     queryset = LiquidityGapResultsBase.objects.all()

#     # Apply filters if they are provided
#     fic_mis_date = None
#     if form.is_valid():
#         process_name = form.cleaned_data.get('process_name')
#         fic_mis_date = form.cleaned_data.get('fic_mis_date')
#         v_ccy_code = form.cleaned_data.get('v_ccy_code')
#         account_type = form.cleaned_data.get('account_type')
#         bucket_number = form.cleaned_data.get('bucket_number')

#         if process_name:
#             queryset = queryset.filter(process_name=process_name)
#         if fic_mis_date:
#             queryset = queryset.filter(fic_mis_date=fic_mis_date)
#         if v_ccy_code:
#             queryset = queryset.filter(v_ccy_code=v_ccy_code)
#         if account_type:
#             queryset = queryset.filter(account_type=account_type)
#         if bucket_number:
#             queryset = queryset.filter(bucket_number=bucket_number)

#     # Fallback to the latest fic_mis_date in Dim_Dates if fic_mis_date is not set
#     if not fic_mis_date:
#         try:
#             fic_mis_date = Dim_Dates.objects.latest('fic_mis_date').fic_mis_date
#         except Dim_Dates.DoesNotExist:
#             messages.error(request, "No data available for the selected filters.")
#             return render(request, 'ALM_APP/reports/liquidity_gap_report.html', {'form': form})

#     # Filter Dim_Dates by fic_mis_date for date buckets
#     date_buckets = Dim_Dates.objects.filter(fic_mis_date=fic_mis_date).values(
#         'bucket_number', 'bucket_start_date', 'bucket_end_date'
#     ).distinct().order_by('bucket_number')

#     # Check if no date buckets found
#     if not date_buckets.exists():
#         messages.error(request, "No date buckets available for the selected filters.")
#         return render(request, 'ALM_APP/reports/liquidity_gap_report.html', {'form': form})

#     # Further filter LiquidityGapResultsBase by fic_mis_date if it's not None
#     queryset = queryset.filter(fic_mis_date=fic_mis_date)

#     # Separate inflows and outflows, grouped by product type and bucket number
#     inflow_products = queryset.filter(account_type="Inflow").values('v_prod_type', 'bucket_number').annotate(total=Sum('inflows'))
#     outflow_products = queryset.filter(account_type="Outflow").values('v_prod_type', 'bucket_number').annotate(total=Sum('outflows'))

#     # Prepare data structure for inflows and outflows
#     inflow_data = {}
#     outflow_data = {}

#     for item in inflow_products:
#         prod_type = item['v_prod_type']
#         bucket = item['bucket_number']
#         total = item['total']
#         if prod_type not in inflow_data:
#             inflow_data[prod_type] = {}
#         inflow_data[prod_type][bucket] = total

#     for item in outflow_products:
#         prod_type = item['v_prod_type']
#         bucket = item['bucket_number']
#         total = item['total']
#         if prod_type not in outflow_data:
#             outflow_data[prod_type] = {}
#         outflow_data[prod_type][bucket] = total

#     # Get the first item of inflow and outflow data for easy access in the template
#     first_inflow_product, remaining_inflow_data = None, inflow_data.copy()
#     first_outflow_product, remaining_outflow_data = None, outflow_data.copy()

#     if inflow_data:
#         first_inflow_product = list(inflow_data.items())[0]
#         remaining_inflow_data.pop(first_inflow_product[0], None)

#     if outflow_data:
#         first_outflow_product = list(outflow_data.items())[0]
#         remaining_outflow_data.pop(first_outflow_product[0], None)

#     # Calculate total inflows and total outflows by bucket
#     total_inflows_by_bucket = {}
#     total_outflows_by_bucket = {}

#     for bucket in date_buckets:
#         bucket_number = bucket['bucket_number']
#         total_inflows_by_bucket[bucket_number] = sum(inflow_data.get(prod, {}).get(bucket_number, 0) for prod in inflow_data)
#         total_outflows_by_bucket[bucket_number] = sum(outflow_data.get(prod, {}).get(bucket_number, 0) for prod in outflow_data)

#     # Calculate Net Liquidity Gap, Net Gap as % of Total Outflows, and Cumulative Gap
#     net_liquidity_gap = {bucket: total_inflows_by_bucket[bucket] - total_outflows_by_bucket[bucket] for bucket in total_inflows_by_bucket}
#     net_gap_percentage = {bucket: (net_liquidity_gap[bucket] / total_outflows_by_bucket[bucket] * 100) if total_outflows_by_bucket[bucket] else 0 for bucket in total_outflows_by_bucket}

#     cumulative_gap = {}
#     cumulative_total = 0
#     for bucket in date_buckets:
#         bucket_number = bucket['bucket_number']
#         cumulative_total += net_liquidity_gap[bucket_number]
#         cumulative_gap[bucket_number] = cumulative_total

#     # Calculate total columns (date buckets + 2 for "Account Type" and "Product")
#     total_columns = len(date_buckets) + 2

#     # Prepare context for rendering in the template
#     context = {
#         'form': form,
#         'fic_mis_date': fic_mis_date,
#         'date_buckets': date_buckets,
#         'first_inflow_product': first_inflow_product,
#         'remaining_inflow_data': remaining_inflow_data,
#         'first_outflow_product': first_outflow_product,
#         'remaining_outflow_data': remaining_outflow_data,
#         'currency_code': "BWP",  # Example currency code
#         'total_columns': total_columns,  # Pass the total column count to the template
#         'net_liquidity_gap': net_liquidity_gap,
#         'net_gap_percentage': net_gap_percentage,
#         'cumulative_gap': cumulative_gap,
#     }

#     return render(request, 'ALM_APP/reports/liquidity_gap_report.html', context)


# View to project cash flows based on the fic_mis_date parameter
@login_required
def project_cash_flows_view(request):
    process_name = 'Blessmoe'
    fic_mis_date = '2024-08-31'
    # status = populate_dim_dates_from_time_buckets(fic_mis_date)
    # status=populate_dim_product(fic_mis_date)
    # status= aggregate_by_prod_code(fic_mis_date, process_name)
    # status=update_date(fic_mis_date)
    # status = populate_liquidity_gap_results_base(fic_mis_date, process_name)
    # status= calculate_time_buckets_and_spread(process_name, fic_mis_date)
    # status= aggregate_cashflows_to_product_level(fic_mis_date)

    

    status= project_cash_flows(fic_mis_date)



    print(status)
    # project_cash_flows(fic_mis_date)
    return render(request, 'ALM_APP/project_cash_flows.html')

@login_required
def dashboard_view(request):
    # Example data for financial graphs
    mis_date = '2024-07-31'  # Input date in 'YYYY-MM-DD' format
    # perform_interpolation(mis_date)
    # project_cash_flows(mis_date)
    # update_cash_flows_with_ead(mis_date)
    # # #Insert records into FCT_Stage_Determination with the numeric date
    # insert_fct_stage(mis_date)
    # # #determine stage
    # update_stage(mis_date)
    # process_cooling_period_for_accounts(mis_date)
    # update_stage_determination(mis_date)
    # update_stage_determination_accrued_interest_and_ead(mis_date)
    # update_stage_determination_eir(mis_date)
    # update_lgd_for_stage_determination_term_structure(mis_date)
    # calculate_pd_for_accounts(mis_date)
    # insert_cash_flow_data(mis_date)
    # update_financial_cash_flow(mis_date)
    # update_cash_flow_with_pd_buckets(mis_date)
    # update_marginal_pd(mis_date)
    # calculate_expected_cash_flow(mis_date)
    # calculate_discount_factors(mis_date)
    # calculate_cashflow_fields(mis_date)
    # calculate_forward_loss_fields(mis_date)
    # populate_fct_reporting_lines(mis_date)
    # calculate_ecl_based_on_method(mis_date)
    # update_reporting_lines_with_exchange_rate(mis_date)

    return render(request, 'ALM_home.html')


@login_required
def ALM_home_view(request):
    context = {
        'title': ' Home',
        # You can pass any additional context if needed
    }
    return render(request, 'ALM_home.html', context)
